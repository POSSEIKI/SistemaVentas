import io
import csv
from decimal import Decimal
from typing import List, Optional
import openpyxl
from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_, func
from sqlalchemy.orm import joinedload

import math
from app.db.database import get_db
from app.core.deps import get_current_user
from app.models.producto import Producto, Categoria, UnidadMedida
from app.schemas.ventas import ProductoCreate, ProductoUpdate, ProductoOut, PaginatedProductosOut
from app.services.inventario_service import generar_excel_inventario_fisico, procesar_ajuste_inventario_fisico

router = APIRouter(prefix="/productos", tags=["Productos"])

def _to_out(p: Producto) -> ProductoOut:
    return ProductoOut(
        id=p.id,
        codigo=p.codigo,
        codigo_barras=p.codigo_barras,
        codigo_barras_blister=p.codigo_barras_blister,
        codigo_barras_unidad=p.codigo_barras_unidad,
        nombre=p.nombre,
        descripcion=p.descripcion,
        precio_venta=p.precio_venta or Decimal("0"),
        precio_costo=p.precio_costo or Decimal("0"),
        iva_porcentaje=p.iva_porcentaje or Decimal("0"),
        stock_actual=p.stock_actual or Decimal("0"),
        stock_minimo=p.stock_minimo or Decimal("0"),
        afecta_inventario=p.afecta_inventario if p.afecta_inventario is not None else True,
        es_servicio=p.es_servicio if p.es_servicio is not None else False,
        activo=p.activo if p.activo is not None else True,
        categoria_id=p.categoria_id,
        categoria_nombre=p.categoria.nombre if p.categoria else None,
        unidad_medida_id=p.unidad_medida_id,
        unidad_abreviatura=p.unidad_medida.abreviatura if p.unidad_medida else None,
        maneja_fracciones=p.maneja_fracciones if p.maneja_fracciones is not None else False,
        contenido_caja=p.contenido_caja or 1,
        contenido_blister=p.contenido_blister or 0,
        precio_caja=p.precio_caja or Decimal("0"),
        precio_blister=p.precio_blister or Decimal("0"),
        precio_unidad=p.precio_unidad or Decimal("0"),
        laboratorio=p.laboratorio,
        principio_activo=p.principio_activo,
        ubicacion=p.ubicacion,
    )

@router.get("/buscar", response_model=List[ProductoOut])
async def buscar_productos(
    q: str = Query(..., min_length=1),
    categoria_id: Optional[int] = None,
    modo: Optional[str] = Query("NOMBRE", enum=["NOMBRE", "SUSTANCIA", "CODIGO"]),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    from sqlalchemy import case
    q_clean = q.strip()
    empresa_id = current_user.empresa_id or 1

    condiciones = [
        Producto.empresa_id == empresa_id,
        Producto.activo == True
    ]

    if modo == "SUSTANCIA":
        # Búsqueda dedicada por principio activo / sustancia genérica
        rank_expr = case(
            (Producto.principio_activo == q_clean, 1),
            (Producto.principio_activo.ilike(f"{q_clean}%"), 2),
            (Producto.principio_activo.ilike(f"% {q_clean}%"), 3),
            (Producto.principio_activo.ilike(f"%{q_clean}%"), 4),
            else_=5
        )
        condiciones.extend([
            Producto.principio_activo.isnot(None),
            Producto.principio_activo.ilike(f"%{q_clean}%")
        ])
    elif modo == "CODIGO":
        # Búsqueda dedicada por código o códigos de barra (caja, blister, unidad)
        rank_expr = case(
            (Producto.codigo_barras == q_clean, 1),
            (Producto.codigo_barras_blister == q_clean, 2),
            (Producto.codigo_barras_unidad == q_clean, 3),
            (Producto.codigo == q_clean, 4),
            (Producto.codigo_barras.ilike(f"{q_clean}%"), 5),
            (Producto.codigo.ilike(f"{q_clean}%"), 6),
            else_=7
        )
        condiciones.append(
            or_(
                Producto.codigo_barras == q_clean,
                Producto.codigo_barras_blister == q_clean,
                Producto.codigo_barras_unidad == q_clean,
                Producto.codigo.ilike(f"%{q_clean}%"),
                Producto.codigo_barras.ilike(f"%{q_clean}%"),
                Producto.codigo_barras_blister.ilike(f"%{q_clean}%"),
                Producto.codigo_barras_unidad.ilike(f"%{q_clean}%"),
            )
        )
    else:
        # Modo NOMBRE (Predeterminado): Prioridad a códigos exactos de barra/código y nombres
        rank_expr = case(
            (Producto.codigo_barras == q_clean, 1),
            (Producto.codigo_barras_blister == q_clean, 2),
            (Producto.codigo_barras_unidad == q_clean, 3),
            (Producto.codigo == q_clean, 4),
            (Producto.nombre.ilike(f"{q_clean}%"), 5),
            (Producto.nombre.ilike(f"% {q_clean}%"), 6),
            (Producto.nombre.ilike(f"%{q_clean}%"), 7),
            (Producto.codigo.ilike(f"{q_clean}%"), 8),
            (Producto.principio_activo.ilike(f"{q_clean}%"), 9),
            (Producto.principio_activo.ilike(f"%{q_clean}%"), 10),
            (Producto.laboratorio.ilike(f"%{q_clean}%"), 11),
            else_=12
        )
        condiciones.append(
            or_(
                Producto.nombre.ilike(f"%{q_clean}%"),
                Producto.codigo.ilike(f"%{q_clean}%"),
                Producto.codigo_barras.ilike(f"%{q_clean}%"),
                Producto.codigo_barras_blister.ilike(f"%{q_clean}%"),
                Producto.codigo_barras_unidad.ilike(f"%{q_clean}%"),
                Producto.principio_activo.ilike(f"%{q_clean}%"),
                Producto.laboratorio.ilike(f"%{q_clean}%"),
            )
        )

    if categoria_id:
        condiciones.append(Producto.categoria_id == categoria_id)

    stmt = (
        select(Producto)
        .options(joinedload(Producto.categoria), joinedload(Producto.unidad_medida))
        .where(*condiciones)
        .order_by(rank_expr, Producto.nombre)
        .limit(50)
    )
    result = await db.execute(stmt)
    productos = result.scalars().unique().all()
    return [_to_out(p) for p in productos]

CATEGORIAS_POR_RUBRO = {
    "FARMACIA": {
        "medicamentos", "analgesicos", "antiinflamatorios", "antibioticos", "vitaminas",
        "suplementos", "cuidado personal", "aseo", "dispositivos medicos", "maternidad",
        "bebes", "primeros auxilios", "drogueria", "farmacia", "general"
    },
    "DROGUERIA": {
        "medicamentos", "analgesicos", "antiinflamatorios", "antibioticos", "vitaminas",
        "suplementos", "cuidado personal", "aseo", "dispositivos medicos", "maternidad",
        "bebes", "primeros auxilios", "drogueria", "farmacia", "general"
    },
    "FERRETERIA": {
        "herramientas", "herramientas manuales", "herramientas electricas", "construccion",
        "tornilleria", "fijaciones", "pinturas", "quimicos", "plomeria", "fontaneria",
        "electricos", "iluminacion", "cerrajeria", "seguridad", "ferreteria", "general"
    },
    "MINIMARKET": {
        "abarrotes", "bebidas", "licores", "lacteos", "huevos", "frutas", "verduras",
        "carnes", "embutidos", "limpieza", "hogar", "snacks", "dulces", "panaderia",
        "minimarket", "supermercado", "viveres", "general"
    },
    "SUPERMERCADO": {
        "abarrotes", "bebidas", "licores", "lacteos", "huevos", "frutas", "verduras",
        "carnes", "embutidos", "limpieza", "hogar", "snacks", "dulces", "panaderia",
        "minimarket", "supermercado", "viveres", "general"
    },
    "RESTAURANTE": {
        "platos a la carta", "bebidas", "refrescos", "desayunos", "postres", "dulces",
        "combos", "promociones", "entradas", "almuerzos", "comidas rapidas", "restaurante", "general"
    },
    "PANADERIA": {
        "panes", "reposteria", "bebidas", "cafeteria", "postres", "dulces", "lacteos",
        "insumos de panaderia", "panaderia", "pasteleria", "general"
    },
    "ROPA": {
        "ropa hombre", "ropa mujer", "ropa infantil", "calzado", "accesorios",
        "ropa deportiva", "moda", "textil", "general"
    },
    "COMERCIO_GENERAL": {
        "general", "articulos varios", "accesorios", "servicios"
    }
}

def _es_categoria_afin(nombre_cat: str, rubro: str) -> bool:
    import unicodedata
    s = str(nombre_cat or "").strip().lower()
    norm = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    if "general" in norm:
        return True
    afines = CATEGORIAS_POR_RUBRO.get(rubro.upper(), CATEGORIAS_POR_RUBRO.get("COMERCIO_GENERAL", set()))
    return any(p in norm for p in afines)

@router.get("/categorias/lista")
async def listar_categorias(
    solo_con_productos: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user)
):
    empresa_id = current_user.empresa_id or 1
    from app.models.configuracion import ConfiguracionEmpresa
    res_cfg = await db.execute(select(ConfiguracionEmpresa).where(ConfiguracionEmpresa.empresa_id == empresa_id).order_by(ConfiguracionEmpresa.id.desc()))
    cfg = res_cfg.scalars().first()
    if not cfg and empresa_id == 1:
        res_cfg = await db.execute(select(ConfiguracionEmpresa).where(ConfiguracionEmpresa.id == 1))
        cfg = res_cfg.scalars().first()
    rubro = (cfg.rubro if cfg and cfg.rubro else "COMERCIO_GENERAL").upper()

    stmt = (
        select(Categoria.id, Categoria.nombre, func.count(Producto.id).label("total_productos"))
        .outerjoin(
            Producto,
            (Producto.categoria_id == Categoria.id) & (Producto.activo == True) & (Producto.empresa_id == empresa_id)
        )
        .where(Categoria.activo == True)
        .group_by(Categoria.id, Categoria.nombre)
        .order_by(Categoria.nombre)
    )
    result = await db.execute(stmt)
    todas = result.all()

    categorias_filtradas = []
    for row in todas:
        # 1. Si la categoría tiene al menos 1 producto registrado en este negocio, SIEMPRE se muestra
        if row.total_productos > 0:
            categorias_filtradas.append({"id": row.id, "nombre": row.nombre, "total_productos": row.total_productos})
        elif not solo_con_productos:
            # 2. Si NO tiene productos, SOLO se muestra si es afín al rubro del negocio
            if _es_categoria_afin(row.nombre, rubro):
                categorias_filtradas.append({"id": row.id, "nombre": row.nombre, "total_productos": 0})

    if not categorias_filtradas:
        for row in todas:
            if "general" in row.nombre.lower():
                categorias_filtradas.append({"id": row.id, "nombre": row.nombre, "total_productos": row.total_productos})
        if not categorias_filtradas:
            categorias_filtradas = [{"id": row.id, "nombre": row.nombre, "total_productos": row.total_productos} for row in todas]

    return categorias_filtradas

@router.post("/categorias")
async def crear_categoria(datos: dict, db: AsyncSession = Depends(get_db), _=Depends(get_current_user)):
    nombre = (datos.get("nombre") or "").strip()
    if not nombre:
        raise HTTPException(status_code=400, detail="Nombre de categoría requerido")
    
    result = await db.execute(select(Categoria).where(Categoria.nombre.ilike(nombre)))
    cat = result.scalar_one_or_none()
    if not cat:
        cat = Categoria(nombre=nombre)
        db.add(cat)
        await db.commit()
        await db.refresh(cat)
    return {"id": cat.id, "nombre": cat.nombre, "total_productos": 0}

@router.get("/unidades/lista")
async def listar_unidades(db: AsyncSession = Depends(get_db), _=Depends(get_current_user)):
    result = await db.execute(select(UnidadMedida).where(UnidadMedida.activo == True).order_by(UnidadMedida.nombre))
    return [{"id": u.id, "nombre": u.nombre, "abreviatura": u.abreviatura} for u in result.scalars().all()]

@router.get("/por-codigo/{codigo}")
async def obtener_por_codigo(codigo: str, db: AsyncSession = Depends(get_db), current_user=Depends(get_current_user)):
    cod_clean = codigo.strip()
    empresa_id = current_user.empresa_id or 1
    stmt = (
        select(Producto)
        .options(joinedload(Producto.categoria), joinedload(Producto.unidad_medida))
        .where(
            Producto.empresa_id == empresa_id,
            or_(
                Producto.codigo == cod_clean,
                Producto.codigo_barras == cod_clean,
                Producto.codigo_barras_blister == cod_clean,
                Producto.codigo_barras_unidad == cod_clean,
            )
        )
    )
    result = await db.execute(stmt)
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    
    # Detectar cuál presentación coincide con el código de barras escaneado
    presentacion = "CAJA"
    if p.maneja_fracciones:
        if p.codigo_barras_blister and p.codigo_barras_blister == cod_clean:
            presentacion = "BLISTER"
        elif p.codigo_barras_unidad and p.codigo_barras_unidad == cod_clean:
            presentacion = "UNIDAD"
        elif p.codigo_barras and p.codigo_barras == cod_clean:
            presentacion = "CAJA"

    return {
        "producto": _to_out(p),
        "presentacion_detectada": presentacion
    }

@router.get("", response_model=PaginatedProductosOut)
async def listar_productos(
    categoria_id: Optional[int] = None,
    q: Optional[str] = None,
    activo: Optional[bool] = True,
    filtro_stock: Optional[str] = Query(None, enum=["TODOS", "CON_STOCK", "SIN_STOCK", "STOCK_BAJO"]),
    pagina: int = Query(1, ge=1),
    limite: int = Query(50, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    empresa_id = current_user.empresa_id or 1
    condiciones = [Producto.empresa_id == empresa_id]
    if activo is not None:
        condiciones.append(Producto.activo == activo)
    if categoria_id:
        condiciones.append(Producto.categoria_id == categoria_id)
    if filtro_stock == "CON_STOCK":
        condiciones.append(Producto.stock_actual > 0)
    elif filtro_stock == "SIN_STOCK":
        condiciones.append(Producto.stock_actual <= 0)
    elif filtro_stock == "STOCK_BAJO":
        condiciones.append(Producto.stock_actual <= Producto.stock_minimo)

    if q and q.strip():
        condiciones.append(
            or_(
                Producto.nombre.ilike(f"%{q.strip()}%"),
                Producto.codigo.ilike(f"%{q.strip()}%"),
                Producto.codigo_barras.ilike(f"%{q.strip()}%"),
                Producto.codigo_barras_blister.ilike(f"%{q.strip()}%"),
                Producto.codigo_barras_unidad.ilike(f"%{q.strip()}%"),
                Producto.principio_activo.ilike(f"%{q.strip()}%"),
                Producto.laboratorio.ilike(f"%{q.strip()}%"),
            )
        )

    # 1. Conteo total para paginación
    count_stmt = select(func.count(Producto.id))
    if condiciones:
        count_stmt = count_stmt.where(*condiciones)
    res_count = await db.execute(count_stmt)
    total = res_count.scalar() or 0

    # 2. Consulta paginada
    offset = (pagina - 1) * limite
    query = (
        select(Producto)
        .options(joinedload(Producto.categoria), joinedload(Producto.unidad_medida))
    )
    if condiciones:
        query = query.where(*condiciones)

    result = await db.execute(query.order_by(Producto.id.desc()).offset(offset).limit(limite))
    productos = result.scalars().unique().all()
    total_paginas = math.ceil(total / limite) if total > 0 else 1

    return PaginatedProductosOut(
        items=[_to_out(p) for p in productos],
        total=total,
        pagina=pagina,
        limite=limite,
        total_paginas=total_paginas,
    )

# ─── EXPORTACIÓN PARA TOMA DE INVENTARIO FÍSICO ────────────────────────────────

@router.get("/exportar-inventario-fisico")
async def exportar_inventario_fisico(
    categoria_id: Optional[int] = None,
    q: Optional[str] = None,
    solo_con_stock: Optional[bool] = False,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    empresa_id = current_user.empresa_id or 1
    excel_stream = await generar_excel_inventario_fisico(
        db=db,
        categoria_id=categoria_id,
        q=q,
        solo_con_stock=solo_con_stock,
        empresa_id=empresa_id,
    )
    from datetime import datetime
    fecha_slug = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"Inventario_Fisico_Toma_{fecha_slug}.xlsx"

    return StreamingResponse(
        excel_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ─── CARGUE Y AJUSTE DE INVENTARIO FÍSICO (DESFASE & CONCILIACIÓN) ───────────

@router.post("/ajustar-inventario-fisico")
async def ajustar_inventario_fisico(
    archivo: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    empresa_id = current_user.empresa_id or 1
    contenido = await archivo.read()
    try:
        resultado = await procesar_ajuste_inventario_fisico(
            contenido_bytes=contenido,
            db=db,
            empresa_id=empresa_id,
        )
        return resultado
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error ajustando inventario físico: {str(e)}")

@router.post("", response_model=ProductoOut)
async def crear_producto(
    datos: ProductoCreate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    empresa_id = current_user.empresa_id or 1
    codigo_clean = (datos.codigo or "").strip()
    nombre_clean = (datos.nombre or "").strip()

    if not codigo_clean:
        raise HTTPException(status_code=400, detail="El código de referencia es obligatorio")
    if not nombre_clean:
        raise HTTPException(status_code=400, detail="El nombre del producto es obligatorio")

    # Validar código único dentro de la misma empresa
    result = await db.execute(
        select(Producto).where(
            Producto.empresa_id == empresa_id,
            Producto.codigo == codigo_clean
        )
    )
    existente = result.scalar_one_or_none()
    if existente:
        raise HTTPException(
            status_code=400,
            detail=f"Ya existe un producto en tu negocio con el código '{codigo_clean}' (ID: {existente.id}). Búscalo en la tabla para modificarlo o usa otro código."
        )

    p_data = datos.model_dump()
    p_data["codigo"] = codigo_clean
    p_data["nombre"] = nombre_clean

    # Limpiar strings vacíos a None para evitar colisiones
    for field in ["codigo_barras", "codigo_barras_blister", "codigo_barras_unidad", "laboratorio", "principio_activo", "ubicacion", "descripcion"]:
        if field in p_data and isinstance(p_data[field], str):
            val_clean = p_data[field].strip()
            p_data[field] = val_clean if val_clean else None

    # Validar que unidad_medida_id exista en la BD para evitar IntegrityError
    if p_data.get("unidad_medida_id"):
        res_u = await db.execute(select(UnidadMedida).where(UnidadMedida.id == p_data["unidad_medida_id"]))
        if not res_u.scalar_one_or_none():
            res_u_first = await db.execute(select(UnidadMedida).order_by(UnidadMedida.id.asc()).limit(1))
            u_first = res_u_first.scalar_one_or_none()
            p_data["unidad_medida_id"] = u_first.id if u_first else None
    else:
        res_u_first = await db.execute(select(UnidadMedida).order_by(UnidadMedida.id.asc()).limit(1))
        u_first = res_u_first.scalar_one_or_none()
        p_data["unidad_medida_id"] = u_first.id if u_first else None

    # Validar que categoria_id exista en la BD si se envió
    if p_data.get("categoria_id"):
        res_cat = await db.execute(select(Categoria).where(Categoria.id == p_data["categoria_id"]))
        if not res_cat.scalar_one_or_none():
            p_data["categoria_id"] = None

    # Si es fraccionado y precio_venta no fue puesto, usar precio_caja o precio_unidad
    if p_data.get("maneja_fracciones"):
        if not p_data.get("precio_venta") and p_data.get("precio_caja"):
            p_data["precio_venta"] = p_data["precio_caja"]

    p_data["empresa_id"] = empresa_id
    res_max_p = await db.execute(select(func.coalesce(func.max(Producto.id), 0)))
    p_data["id"] = (res_max_p.scalar() or 0) + 1
    producto = Producto(**p_data)
    db.add(producto)
    await db.commit()
    
    # Reload with relationships
    stmt = (
        select(Producto)
        .options(joinedload(Producto.categoria), joinedload(Producto.unidad_medida))
        .where(Producto.id == producto.id)
    )
    res = await db.execute(stmt)
    return _to_out(res.scalar_one())

@router.patch("/{producto_id}", response_model=ProductoOut)
async def actualizar_producto(
    producto_id: int,
    datos: ProductoUpdate,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    empresa_id = current_user.empresa_id or 1
    result = await db.execute(
        select(Producto).where(
            Producto.id == producto_id,
            Producto.empresa_id == empresa_id
        )
    )
    producto = result.scalar_one_or_none()
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    
    update_data = datos.model_dump(exclude_unset=True)

    # Validar unidad de medida si se intenta actualizar
    if "unidad_medida_id" in update_data and update_data["unidad_medida_id"]:
        res_u = await db.execute(select(UnidadMedida).where(UnidadMedida.id == update_data["unidad_medida_id"]))
        if not res_u.scalar_one_or_none():
            update_data["unidad_medida_id"] = producto.unidad_medida_id

    # Validar categoria si se intenta actualizar
    if "categoria_id" in update_data and update_data["categoria_id"]:
        res_cat = await db.execute(select(Categoria).where(Categoria.id == update_data["categoria_id"]))
        if not res_cat.scalar_one_or_none():
            update_data["categoria_id"] = None

    # Limpiar strings vacíos a None
    for field in ["codigo_barras", "codigo_barras_blister", "codigo_barras_unidad", "laboratorio", "principio_activo", "ubicacion", "descripcion"]:
        if field in update_data and isinstance(update_data[field], str):
            val_clean = update_data[field].strip()
            update_data[field] = val_clean if val_clean else None

    for campo, valor in update_data.items():
        setattr(producto, campo, valor)
    
    await db.commit()
    
    # Reload with relationships
    stmt = (
        select(Producto)
        .options(joinedload(Producto.categoria), joinedload(Producto.unidad_medida))
        .where(Producto.id == producto_id)
    )
    res = await db.execute(stmt)
    return _to_out(res.scalar_one())

@router.delete("/{producto_id}")
async def eliminar_producto(producto_id: int, db: AsyncSession = Depends(get_db), current_user=Depends(get_current_user)):
    empresa_id = current_user.empresa_id or 1
    result = await db.execute(
        select(Producto).where(
            Producto.id == producto_id,
            Producto.empresa_id == empresa_id
        )
    )
    producto = result.scalar_one_or_none()
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    producto.activo = False
    await db.commit()
    return {"mensaje": "Producto desactivado"}

# ─── PLANTILLA DE EXCEL ───────────────────────────────────────────────────────

@router.get("/plantilla-excel")
async def descargar_plantilla_excel(_=Depends(get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = [
        "Codigo", "Codigo_Barras", "Nombre", "Categoria", "Unidad_Medida",
        "Precio_Costo", "Precio_Venta", "IVA_Porcentaje", "Stock_Actual", "Stock_Minimo",
        "Maneja_Fracciones_SI_NO", "Contenido_Caja", "Contenido_Blister",
        "Precio_Caja", "Precio_Blister", "Precio_Unidad", "Laboratorio_Marca", "Principio_Activo", "Ubicacion"
    ]
    ws.append(headers)

    # Filas de ejemplo
    ejemplos = [
        [
            "100026176", "7702057001234", "ACETAMINOFEN 500 MG 100 TAB MK", "Medicamentos", "CAJA",
            12000, 18000, 0, 150, 20,
            "SI", 100, 10,
            18000, 2000, 300, "TECNOQUIMICAS", "ACETAMINOFEN", "Estante A1"
        ],
        [
            "670", "", "ABRAZADERA 1 PULGADA TITAN", "Ferreteria", "UNIDAD",
            1200, 2000, 19, 50, 10,
            "NO", 1, 0,
            2000, 0, 0, "TITAN", "", "Bodega B"
        ]
    ]
    for ej in ejemplos:
        ws.append(ej)

    # Estilo básico
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_productos_sistemaventas.xlsx"}
    )

# ─── IMPORTADOR INTELIGENTE (EXCEL / CSV) ──────────────────────────────────────

@router.post("/importar-excel")
async def importar_archivo_productos(
    archivo: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    from app.services.import_service import procesar_archivo_inventario
    contenido = await archivo.read()
    empresa_id = current_user.empresa_id or 1
    try:
        resultado = await procesar_archivo_inventario(
            contenido=contenido,
            nombre_archivo=archivo.filename or "",
            db=db,
            empresa_id=empresa_id,
        )
        return resultado
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error procesando archivo: {str(e)}")

# ─── APLICAR REDONDEO GLOBAL A TODO EL CATÁLOGO ───────────────────────────────

@router.post("/aplicar-redondeo-global")
async def aplicar_redondeo_global(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    from app.models.configuracion import ConfiguracionEmpresa
    empresa_id = current_user.empresa_id or 1
    res_cfg = await db.execute(
        select(ConfiguracionEmpresa).where(ConfiguracionEmpresa.empresa_id == empresa_id).order_by(ConfiguracionEmpresa.id.desc())
    )
    cfg = res_cfg.scalars().first()
    if not cfg and empresa_id == 1:
        res_cfg = await db.execute(select(ConfiguracionEmpresa).where(ConfiguracionEmpresa.id == 1))
        cfg = res_cfg.scalars().first()
    modo = getattr(cfg, "modo_redondeo", "CENTENA_100") or "CENTENA_100"

    def _redondear(val: float, modo_r: str) -> float:
        if val <= 0:
            return 0.0
        if modo_r == "ENTERO":
            return float(round(val))
        elif modo_r == "CINCUENTA_50":
            return float(round(val / 50.0) * 50)
        elif modo_r == "CENTENA_100":
            return float(round(val / 100.0) * 100)
        elif modo_r == "MIL_1000":
            return float(round(val / 1000.0) * 1000)
        elif modo_r == "DECIMALES_2":
            return round(val, 2)
        else:
            return float(round(val / 100.0) * 100)

    res_prods = await db.execute(select(Producto).where(Producto.empresa_id == empresa_id))
    prods = res_prods.scalars().all()
    modificados = 0
    for p in prods:
        changed = False
        try:
            if p.precio_venta is not None:
                v = float(p.precio_venta)
                if v > 0:
                    new_v = Decimal(str(_redondear(v, modo)))
                    if p.precio_venta != new_v:
                        p.precio_venta = new_v
                        changed = True
            if p.precio_caja is not None:
                c = float(p.precio_caja)
                if c > 0:
                    new_c = Decimal(str(_redondear(c, modo)))
                    if p.precio_caja != new_c:
                        p.precio_caja = new_c
                        changed = True
            if p.precio_blister is not None:
                b = float(p.precio_blister)
                if b > 0:
                    new_b = Decimal(str(_redondear(b, modo)))
                    if p.precio_blister != new_b:
                        p.precio_blister = new_b
                        changed = True
            if p.precio_unidad is not None:
                u = float(p.precio_unidad)
                if u > 0:
                    new_u = Decimal(str(_redondear(u, modo)))
                    if p.precio_unidad != new_u:
                        p.precio_unidad = new_u
                        changed = True
        except Exception:
            continue
        if changed:
            modificados += 1

    await db.commit()
    return {
        "mensaje": f"Se aplicó el redondeo ({modo}) a {modificados} productos del catálogo",
        "productos_actualizados": modificados,
        "modo_aplicado": modo,
    }

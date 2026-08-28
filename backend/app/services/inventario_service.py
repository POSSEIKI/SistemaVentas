import io
import re
import math
import unicodedata
from decimal import Decimal
from datetime import datetime
from typing import Optional, List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_, func
from sqlalchemy.orm import joinedload

from app.models.producto import Producto, Categoria
from app.models.configuracion import ConfiguracionEmpresa

def _normalizar(texto: Any) -> str:
    if texto is None:
        return ""
    s = str(texto).strip().lower()
    return ''.join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )

async def generar_excel_inventario_fisico(
    db: AsyncSession,
    categoria_id: Optional[int] = None,
    q: Optional[str] = None,
    solo_con_stock: bool = False,
) -> io.BytesIO:
    res_cfg = await db.execute(select(ConfiguracionEmpresa).where(ConfiguracionEmpresa.id == 1))
    cfg = res_cfg.scalar_one_or_none()
    rubro = (cfg.rubro if cfg and cfg.rubro else "FARMACIA").upper()
    nombre_empresa = cfg.nombre if cfg and cfg.nombre else "SistemaVentas"

    stmt = (
        select(Producto)
        .options(joinedload(Producto.categoria), joinedload(Producto.unidad_medida))
        .where(Producto.activo == True)
    )
    if categoria_id:
        stmt = stmt.where(Producto.categoria_id == categoria_id)
    if q and q.strip():
        stmt = stmt.where(
            or_(
                Producto.nombre.ilike(f"%{q.strip()}%"),
                Producto.codigo.ilike(f"%{q.strip()}%"),
                Producto.codigo_barras.ilike(f"%{q.strip()}%"),
                Producto.codigo_barras_blister.ilike(f"%{q.strip()}%"),
                Producto.codigo_barras_unidad.ilike(f"%{q.strip()}%"),
                Producto.principio_activo.ilike(f"%{q.strip()}%"),
            )
        )
    if solo_con_stock:
        stmt = stmt.where(Producto.stock_actual > 0)

    stmt = stmt.order_by(Producto.categoria_id, Producto.nombre)
    result = await db.execute(stmt)
    productos = result.scalars().unique().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Fisico"

    font_titulo = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    font_subtitulo = Font(name="Calibri", size=10, italic=True, color="E2E8F0")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=10)
    font_conteo = Font(name="Calibri", size=10, bold=True, color="1E3A8A")

    fill_header_main = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_header_conteo = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    fill_header_desfase = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    fill_conteo_cell = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    thick_conteo_border = Border(
        left=Side(style="medium", color="D97706"),
        right=Side(style="medium", color="D97706"),
        top=Side(style="thin", color="D97706"),
        bottom=Side(style="thin", color="D97706"),
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    title_cell.value = f"HOJA DE TOMA DE INVENTARIO FÍSICO — {nombre_empresa.upper()}"
    title_cell.font = font_titulo
    title_cell.fill = fill_header_main
    title_cell.alignment = align_center

    ws.merge_cells("A2:M2")
    sub_cell = ws["A2"]
    fecha_hoy = datetime.now().strftime("%Y-%m-%d %H:%M")
    sub_cell.value = f"Generado: {fecha_hoy} · Rubro: {rubro} · Total Productos: {len(productos)} · Diligencie el conteo físico en las columnas amarillas"
    sub_cell.font = font_subtitulo
    sub_cell.fill = fill_header_main
    sub_cell.alignment = align_center

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 10

    if rubro == "FARMACIA":
        headers = [
            ("ID", "left", fill_header_main),
            ("CODIGO", "left", fill_header_main),
            ("COD_BARRAS_CAJA", "left", fill_header_main),
            ("COD_BARRAS_BLISTER", "left", fill_header_main),
            ("COD_BARRAS_UNIDAD", "left", fill_header_main),
            ("NOMBRE_COMERCIAL", "left", fill_header_main),
            ("CATEGORIA", "left", fill_header_main),
            ("LABORATORIO", "left", fill_header_main),
            ("PRINCIPIO_ACTIVO", "left", fill_header_main),
            ("UBICACION", "left", fill_header_main),
            ("FRACCION", "center", fill_header_main),
            ("STOCK_DIGITAL_TOTAL", "right", fill_header_main),
            ("CONTEO_CAJAS", "center", fill_header_conteo),
            ("CONTEO_BLISTERS", "center", fill_header_conteo),
            ("CONTEO_UNIDADES", "center", fill_header_conteo),
            ("TOTAL_FISICO", "right", fill_header_conteo),
            ("DESFASE", "right", fill_header_desfase),
            ("COSTO_UNIT", "right", fill_header_main),
            ("PRECIO_VENTA", "right", fill_header_main),
        ]
    else:
        headers = [
            ("ID", "left", fill_header_main),
            ("CODIGO", "left", fill_header_main),
            ("COD_BARRAS", "left", fill_header_main),
            ("NOMBRE_PRODUCTO", "left", fill_header_main),
            ("CATEGORIA", "left", fill_header_main),
            ("UBICACION", "left", fill_header_main),
            ("STOCK_DIGITAL", "right", fill_header_main),
            ("CONTEO_FISICO", "center", fill_header_conteo),
            ("DESFASE", "right", fill_header_desfase),
            ("COSTO_UNIT", "right", fill_header_main),
            ("PRECIO_VENTA", "right", fill_header_main),
        ]

    header_row = 4
    ws.row_dimensions[header_row].height = 24

    for col_idx, (h_title, align, fill_color) in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = h_title
        cell.font = font_header
        cell.fill = fill_color
        cell.alignment = align_center
        cell.border = thin_border

    for r_idx, p in enumerate(productos, start=5):
        ws.row_dimensions[r_idx].height = 20
        cat_nom = p.categoria.nombre if p.categoria else ""
        stock_num = float(p.stock_actual or 0)
        costo_num = float(p.precio_costo or 0)
        venta_num = float(p.precio_venta or 0)

        is_even = (r_idx % 2 == 0)
        row_fill = fill_zebra if is_even else PatternFill(fill_type=None)

        if rubro == "FARMACIA":
            c_caja = int(p.contenido_caja or 1)
            c_blister = int(p.contenido_blister or 0)
            unids_blister = int(c_caja / c_blister) if c_blister > 0 else 0
            frac_str = f"Caja x{c_caja}" if p.maneja_fracciones else "No"

            # Columnas: M=Conteo Cajas, N=Conteo Blisters, O=Conteo Unids, P=Total Fisico, Q=Desfase, L=Stock Digital
            formula_total = f"=IF(AND(M{r_idx}=\"\",N{r_idx}=\"\",O{r_idx}=\"\"),\"\",(IF(ISBLANK(M{r_idx}),0,M{r_idx})*{c_caja})+(IF(ISBLANK(N{r_idx}),0,N{r_idx})*{unids_blister})+(IF(ISBLANK(O{r_idx}),0,O{r_idx})))"
            formula_desfase = f"=IF(P{r_idx}=\"\",\"\",P{r_idx}-L{r_idx})"

            row_data = [
                (p.id, align_left, row_fill, thin_border, None),
                (p.codigo, align_left, row_fill, thin_border, None),
                (p.codigo_barras or "", align_left, row_fill, thin_border, None),
                (p.codigo_barras_blister or "", align_left, row_fill, thin_border, None),
                (p.codigo_barras_unidad or "", align_left, row_fill, thin_border, None),
                (p.nombre, align_left, row_fill, thin_border, None),
                (cat_nom, align_left, row_fill, thin_border, None),
                (p.laboratorio or "", align_left, row_fill, thin_border, None),
                (p.principio_activo or "", align_left, row_fill, thin_border, None),
                (p.ubicacion or "", align_left, row_fill, thin_border, None),
                (frac_str, align_center, row_fill, thin_border, None),
                (stock_num, align_right, row_fill, thin_border, "#,##0"),
                ("", align_center, fill_conteo_cell, thick_conteo_border, "#,##0"),
                ("", align_center, fill_conteo_cell, thick_conteo_border, "#,##0"),
                ("", align_center, fill_conteo_cell, thick_conteo_border, "#,##0"),
                (formula_total, align_right, row_fill, thin_border, "#,##0"),
                (formula_desfase, align_right, row_fill, thin_border, "+#,##0;-#,##0;0"),
                (costo_num, align_right, row_fill, thin_border, "$#,##0"),
                (venta_num, align_right, row_fill, thin_border, "$#,##0"),
            ]
        else:
            formula_desfase = f"=IF(H{r_idx}=\"\",\"\",H{r_idx}-G{r_idx})"

            row_data = [
                (p.id, align_left, row_fill, thin_border, None),
                (p.codigo, align_left, row_fill, thin_border, None),
                (p.codigo_barras or "", align_left, row_fill, thin_border, None),
                (p.nombre, align_left, row_fill, thin_border, None),
                (cat_nom, align_left, row_fill, thin_border, None),
                (p.ubicacion or "", align_left, row_fill, thin_border, None),
                (stock_num, align_right, row_fill, thin_border, "#,##0"),
                ("", align_center, fill_conteo_cell, thick_conteo_border, "#,##0"),
                (formula_desfase, align_right, row_fill, thin_border, "+#,##0;-#,##0;0"),
                (costo_num, align_right, row_fill, thin_border, "$#,##0"),
                (venta_num, align_right, row_fill, thin_border, "$#,##0"),
            ]

        for col_idx, (val, alignment, cell_fill, cell_border, num_format) in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.value = val
            cell.font = font_conteo if cell_fill == fill_conteo_cell else font_data
            if cell_fill:
                cell.fill = cell_fill
            cell.alignment = alignment
            cell.border = cell_border
            if num_format:
                cell.number_format = num_format

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]:
                continue
            v_str = str(cell.value or "")
            if cell.number_format and "$" in cell.number_format:
                v_str += "    "
            if len(v_str) > max_len:
                max_len = len(v_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    ws.freeze_panes = "F5"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

async def procesar_ajuste_inventario_fisico(
    contenido_bytes: bytes,
    db: AsyncSession
) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(filename=io.BytesIO(contenido_bytes), data_only=True)
    ws = wb.active

    header_row_idx = None
    headers_map = {}

    for row_idx in range(1, 10):
        row_vals = [_normalizar(ws.cell(row=row_idx, column=c).value) for c in range(1, ws.max_column + 1)]
        if any(k in row_vals for k in ["codigo", "id", "cod", "nombre_comercial", "nombre_producto", "nombre"]):
            header_row_idx = row_idx
            for c_idx, val in enumerate(row_vals, 1):
                if val:
                    headers_map[val] = c_idx
            break

    if not header_row_idx:
        raise ValueError("No se encontró la fila de encabezados en el archivo Excel de conteo físico.")

    col_id = headers_map.get("id") or headers_map.get("id_sistema")
    col_codigo = headers_map.get("codigo") or headers_map.get("cod") or headers_map.get("cod_producto") or headers_map.get("ref")
    col_nombre = headers_map.get("nombre_comercial") or headers_map.get("nombre_producto") or headers_map.get("nombre")

    col_barras_caja = headers_map.get("cod_barras_caja") or headers_map.get("cod_barras") or headers_map.get("codigo_barras") or headers_map.get("barra")
    col_barras_blister = headers_map.get("cod_barras_blister") or headers_map.get("codigo_barras_blister") or headers_map.get("barra_blister")
    col_barras_unidad = headers_map.get("cod_barras_unidad") or headers_map.get("codigo_barras_unidad") or headers_map.get("barra_unidad")

    col_conteo_fisico = headers_map.get("conteo_fisico") or headers_map.get("fisico")
    col_conteo_cajas = headers_map.get("conteo_cajas")
    col_conteo_blisters = headers_map.get("conteo_blisters")
    col_conteo_unidades = headers_map.get("conteo_unidades")
    col_total_fisico = headers_map.get("total_fisico")

    if not col_id and not col_codigo and not col_nombre:
        raise ValueError("El archivo debe contener al menos la columna 'ID', 'CODIGO' o 'NOMBRE' del producto.")

    stmt = select(Producto).options(joinedload(Producto.categoria), joinedload(Producto.unidad_medida)).where(Producto.activo == True)
    res = await db.execute(stmt)
    prods_all = res.scalars().unique().all()
    map_by_id = {p.id: p for p in prods_all}
    map_by_codigo = {p.codigo.strip().upper(): p for p in prods_all if p.codigo}
    map_by_nombre = {_normalizar(p.nombre): p for p in prods_all if p.nombre}

    total_leidos = 0
    total_ajustados = 0
    total_sin_cambio = 0
    barras_actualizadas_count = 0
    sobrantes_count = 0
    faltantes_count = 0
    impacto_total_costo = Decimal("0")
    desfases_detalle = []

    for r in range(header_row_idx + 1, ws.max_row + 1):
        id_val = ws.cell(row=r, column=col_id).value if col_id else None
        cod_val = str(ws.cell(row=r, column=col_codigo).value or "").strip().upper() if col_codigo else None
        nom_val = _normalizar(ws.cell(row=r, column=col_nombre).value) if col_nombre else None

        prod: Optional[Producto] = None
        if id_val is not None:
            try:
                prod = map_by_id.get(int(id_val))
            except:
                pass
        if not prod and cod_val:
            prod = map_by_codigo.get(cod_val)
        if not prod and nom_val:
            prod = map_by_nombre.get(nom_val)

        if not prod:
            continue

        # ── 1. ACTUALIZAR CÓDIGOS DE BARRA SI VIENEN EN EL EXCEL ────────
        modifico_barras = False
        if col_barras_caja:
            b_caja = str(ws.cell(row=r, column=col_barras_caja).value or "").strip()
            if b_caja and b_caja != "None" and b_caja != str(prod.codigo_barras or ""):
                prod.codigo_barras = b_caja
                modifico_barras = True

        if col_barras_blister:
            b_blis = str(ws.cell(row=r, column=col_barras_blister).value or "").strip()
            if b_blis and b_blis != "None" and b_blis != str(prod.codigo_barras_blister or ""):
                prod.codigo_barras_blister = b_blis
                modifico_barras = True

        if col_barras_unidad:
            b_unid = str(ws.cell(row=r, column=col_barras_unidad).value or "").strip()
            if b_unid and b_unid != "None" and b_unid != str(prod.codigo_barras_unidad or ""):
                prod.codigo_barras_unidad = b_unid
                modifico_barras = True

        if modifico_barras:
            barras_actualizadas_count += 1

        # ── 2. CALCULAR CONTEO FÍSICO Y DESFASE ────────────────────────
        conteo_final = None

        if col_total_fisico:
            v = ws.cell(row=r, column=col_total_fisico).value
            if v is not None and str(v).strip() != "":
                try:
                    conteo_final = Decimal(str(v).replace(",", "."))
                except:
                    pass

        if conteo_final is None and (col_conteo_cajas or col_conteo_blisters or col_conteo_unidades):
            val_cajas = ws.cell(row=r, column=col_conteo_cajas).value if col_conteo_cajas else None
            val_blisters = ws.cell(row=r, column=col_conteo_blisters).value if col_conteo_blisters else None
            val_unids = ws.cell(row=r, column=col_conteo_unidades).value if col_conteo_unidades else None

            if any(x is not None and str(x).strip() != "" for x in [val_cajas, val_blisters, val_unids]):
                cajas = Decimal(str(val_cajas or 0).replace(",", ".")) if val_cajas not in [None, ""] else Decimal("0")
                blisters = Decimal(str(val_blisters or 0).replace(",", ".")) if val_blisters not in [None, ""] else Decimal("0")
                unids = Decimal(str(val_unids or 0).replace(",", ".")) if val_unids not in [None, ""] else Decimal("0")

                c_caja = Decimal(str(prod.contenido_caja or 1))
                c_blister = Decimal(str(prod.contenido_blister or 0))
                unids_per_blister = (c_caja / c_blister) if c_blister > 0 else Decimal("0")

                conteo_final = (cajas * c_caja) + (blisters * unids_per_blister) + unids

        if conteo_final is None and col_conteo_fisico:
            v = ws.cell(row=r, column=col_conteo_fisico).value
            if v is not None and str(v).strip() != "":
                try:
                    conteo_final = Decimal(str(v).replace(",", "."))
                except:
                    pass

        if conteo_final is None:
            continue

        total_leidos += 1
        stock_anterior = Decimal(str(prod.stock_actual or 0))
        desfase = conteo_final - stock_anterior

        if desfase != Decimal("0"):
            costo_unit = Decimal(str(prod.precio_costo or 0))
            impacto_item = desfase * costo_unit
            impacto_total_costo += impacto_item

            if desfase > 0:
                sobrantes_count += 1
            else:
                faltantes_count += 1

            desfases_detalle.append({
                "id": prod.id,
                "codigo": prod.codigo,
                "nombre": prod.nombre,
                "categoria": prod.categoria.nombre if prod.categoria else "Sin categoría",
                "stock_digital": float(stock_anterior),
                "conteo_fisico": float(conteo_final),
                "desfase": float(desfase),
                "costo_unitario": float(costo_unit),
                "impacto_costo": float(impacto_item),
                "codigo_barras": prod.codigo_barras,
                "codigo_barras_blister": prod.codigo_barras_blister,
            })

            prod.stock_actual = conteo_final
            total_ajustados += 1
        else:
            total_sin_cambio += 1

    await db.commit()

    return {
        "total_contados": total_leidos,
        "total_ajustados": total_ajustados,
        "total_coincidentes": total_sin_cambio,
        "barras_actualizadas": barras_actualizadas_count,
        "sobrantes": sobrantes_count,
        "faltantes": faltantes_count,
        "impacto_total_costo": float(impacto_total_costo),
        "desfases": desfases_detalle,
    }


async def analizar_factura_compra_excel(
    file_bytes: bytes,
    filename: str,
    db: AsyncSession,
) -> Dict[str, Any]:
    import csv
    import json
    import re
    import html
    import xml.etree.ElementTree as ET

    # 1. Configuración de rubro, margen predeterminado y modo de redondeo
    res_cfg = await db.execute(select(ConfiguracionEmpresa).where(ConfiguracionEmpresa.id == 1))
    cfg = res_cfg.scalar_one_or_none()
    rubro = (cfg.rubro if cfg and cfg.rubro else "FARMACIA").upper()
    margen_def = float(getattr(cfg, "margen_ganancia_predeterminado", 30.0) or 30.0)
    modo_redondeo = getattr(cfg, "modo_redondeo", "CENTENA_100") or "CENTENA_100"

    def _redondear_precio(val: float, modo: str = "CENTENA_100") -> float:
        if val <= 0:
            return 0.0
        if modo == "ENTERO":
            return float(round(val))
        elif modo == "CINCUENTA_50":
            return float(round(val / 50.0) * 50)
        elif modo == "CENTENA_100":
            return float(round(val / 100.0) * 100)
        elif modo == "MIL_1000":
            return float(round(val / 1000.0) * 1000)
        elif modo == "DECIMALES_2":
            return round(val, 2)
        else:
            return float(round(val / 100.0) * 100)

    # 2. Cargar todos los productos para cruce rápido
    res_prods = await db.execute(
        select(Producto)
        .options(joinedload(Producto.categoria), joinedload(Producto.unidad_medida))
        .where(Producto.activo == True)
    )
    productos = res_prods.scalars().unique().all()

    prods_by_id = {p.id: p for p in productos}
    prods_by_codigo = {_normalizar(p.codigo): p for p in productos if p.codigo}
    prods_by_barcode = {}
    for p in productos:
        if p.codigo_barras:
            prods_by_barcode[_normalizar(p.codigo_barras)] = p
        if p.codigo_barras_blister:
            prods_by_barcode[_normalizar(p.codigo_barras_blister)] = p
        if p.codigo_barras_unidad:
            prods_by_barcode[_normalizar(p.codigo_barras_unidad)] = p
    prods_by_nombre = {_normalizar(p.nombre): p for p in productos if p.nombre}

    def _detectar_dialecto_numerico_texto(texto: str) -> str:
        """
        Analiza si el documento/bloque usa:
        - 'LATIN': Punto para miles (100.000, 15.790) y Coma para decimales (100.000,00, 6,3331)
        - 'ANGLO': Coma para miles (100,000.00) y Punto para decimales (6.3331)
        """
        has_latin_both = bool(re.search(r'\b\d{1,3}(?:\.\d{3})+,\d{1,4}\b', texto))
        has_anglo_both = bool(re.search(r'\b\d{1,3}(?:,\d{3})+\.\d{1,4}\b', texto))
        if has_latin_both:
            return "LATIN"
        if has_anglo_both:
            return "ANGLO"
        
        comas_dec = len(re.findall(r'\b\d+,\d{1,4}\b', texto))
        puntos_miles = len(re.findall(r'\b\d{1,3}\.\d{3}(?:\.\d{3})*\b', texto))
        puntos_dec = len(re.findall(r'\b\d+\.\d{1,2}\b', texto))
        
        if comas_dec > 0 or puntos_miles > 0:
            return "LATIN"
        elif puntos_dec > 0 and comas_dec == 0:
            return "ANGLO"
        return "LATIN"

    def _smart_parse_num(val, is_price=True, default=0.0, dialecto="LATIN"):
        if val is None:
            return default
        if isinstance(val, (int, float)):
            return float(val)

        s = str(val).replace("$", "").replace("COP", "").replace("cop", "").replace(" ", "").strip()
        if not s or s.lower() == "none" or s == "-":
            return default

        negativo = s.startswith("-") or s.endswith("-")
        s = s.replace("-", "")

        # Caso 1: Ambos separadores presentes (, y .)
        if "." in s and "," in s:
            last_dot = s.rfind(".")
            last_comma = s.rfind(",")
            if last_comma > last_dot:
                # 100.000,00 o 12.500,50 -> Punto miles, Coma decimal
                s_clean = s[:last_comma].replace(".", "").replace(",", "") + "." + s[last_comma+1:]
            else:
                # 100,000.00 o 12,500.50 -> Coma miles, Punto decimal
                s_clean = s[:last_dot].replace(",", "").replace(".", "") + "." + s[last_dot+1:]

        # Caso 2: Solo contiene comas
        elif "," in s:
            partes = s.split(",")
            if len(partes) > 2:
                # 1,250,000
                s_clean = s.replace(",", "")
            else:
                decimal_part = partes[1]
                if dialecto == "ANGLO" and len(decimal_part) == 3 and is_price and int(partes[0]) > 0:
                    s_clean = s.replace(",", "")
                else:
                    # 6,3331 o 100000,00 -> Coma decimal
                    s_clean = partes[0] + "." + decimal_part

        # Caso 3: Solo contiene puntos
        elif "." in s:
            partes = s.split(".")
            if len(partes) > 2:
                # 1.250.000 -> Miles
                s_clean = s.replace(".", "")
            else:
                decimal_part = partes[1]
                # Si tiene exactamente 3 digitos en la parte decimal:
                # ej 100.000, 15.790, 24.900, 71.700 -> Miles en Colombia
                if len(decimal_part) == 3:
                    if is_price or float(partes[0]) > 50:
                        s_clean = s.replace(".", "")
                    else:
                        s_clean = s
                else:
                    # ej '15790.0', '24900.0', '6.139', '100.0' -> Float estándar
                    s_clean = s
        else:
            s_clean = s

        try:
            num = float(s_clean)
            return -num if negativo else num
        except:
            return default

    def _detectar_promocion_o_regalo(nombre: str, costo: float) -> tuple:
        """
        Detecta si un ítem de factura es un obsequio/bonificación sin precio (OBS)
        o un pre-pack/combo promocional compuesto (ej. 2 DTE, PACK X 2, 2X1, COMBO, GTS).
        Nota: 'DTE' por sí solo significa Desodorante y NO es un combo; solo es combo si inicia con número (ej: '2 DTE').
        Retorna (es_obsequio_probable, es_combo_probable, factor_combo_sugerido).
        """
        nom_upper = (nombre or "").upper().strip()
        
        # 1. Detección de obsequio / bonificación (OBS, BONIF, MUESTRA o costo <= 0)
        # ESTRICTAMENTE RESTRICTIVO / OBLIGATORIO DE CONVERTIR
        es_obs = False
        if costo <= 0:
            es_obs = True
        elif re.search(r"\b(OBS\.?|OBSEQUIO|BONIF\.?|MUESTRA)\b", nom_upper) or nom_upper.startswith("OBS "):
            es_obs = True
            
        # 2. Detección de combo / pre-pack promocional (OPCIONAL DE DESCOMPONER)
        es_comb = False
        factor_sug = 1
        
        # Pre-packs con número antes de DTE (ej: "2 DTE.AXE...", "3 DTE...", "2DTE")
        match_dte = re.search(r"\b(\d+)\s*DTE\b", nom_upper)
        if match_dte:
            try:
                cant = int(match_dte.group(1))
                if cant >= 2:
                    es_comb = True
                    factor_sug = cant
            except:
                pass
            
        # Buscar "PACK X 2", "PACK X 3", "PACKX2"
        match_pack = re.search(r"\bPACK\s*X\s*(\d+)\b", nom_upper)
        if match_pack:
            try:
                cant = int(match_pack.group(1))
                if cant >= 2:
                    es_comb = True
                    factor_sug = cant
            except:
                pass
                
        # Buscar "2X1", "3X2", "2 X 1", "3 X 2"
        if not es_comb:
            match_nxm = re.search(r"\b(\d+)\s*X\s*(\d+)\b", nom_upper)
            if match_nxm:
                try:
                    c1 = int(match_nxm.group(1))
                    if c1 >= 2:
                        es_comb = True
                        factor_sug = c1
                except:
                    pass

        # Buscar "DUO PACK", "TRIO PACK", "DUOPACK", "TRIOPACK", "COMBO", "GTS"
        if not es_comb:
            if re.search(r"\b(DUO\s*PACK|DUOPACK)\b", nom_upper):
                es_comb = True
                factor_sug = 2
            elif re.search(r"\b(TRIO\s*PACK|TRIOPACK)\b", nom_upper):
                es_comb = True
                factor_sug = 3
            elif re.search(r"\bCOMBO\b", nom_upper) and not es_obs:
                es_comb = True
                factor_sug = 2
            elif re.search(r"\b(\d+)\s*UND\s*GTS\b", nom_upper) or re.search(r"\b\+\s*1\s*GTS\b", nom_upper) or re.search(r"\bGRATIS\b", nom_upper):
                es_comb = True
                factor_sug = 2
            
        return es_obs, es_comb, max(1, factor_sug)

    def _parse_pharma_description(desc: str) -> dict:
        if not desc:
            return {"nombre": "", "lote": None, "vencimiento": None, "contenido_caja": 1}
        s = str(desc).strip()
        s = re.sub(r'(\d{4}[-/.]\d{1,2}[-/.])\s+(\d{1,2})', r'\1\2', s)
        match_fecha = re.search(r'\b(20\d{2}[-/.]\d{1,2}[-/.]\d{1,2}|\d{1,2}[-/.]\d{1,2}[-/.]20\d{2})\b', s)
        vencimiento = None
        lote = None
        nombre_limpio = s

        if match_fecha:
            vencimiento = match_fecha.group(1)
            resto = s[:match_fecha.start()].strip()
            tokens = resto.split()
            if tokens:
                posible_lote = tokens[-1]
                if len(posible_lote) >= 3 and any(c.isdigit() for c in posible_lote):
                    lote = posible_lote
                    nombre_limpio = " ".join(tokens[:-1])
                else:
                    nombre_limpio = resto

        contenido_caja = 1
        match_cant = re.search(r'\bX\s*(\d+)\s*(TAB|CAP|COMP|BLIST|SOB|UND)\b', nombre_limpio, re.IGNORECASE)
        if match_cant:
            try:
                contenido_caja = int(match_cant.group(1))
            except:
                pass

        return {
            "nombre": nombre_limpio.strip(),
            "lote": lote,
            "vencimiento": vencimiento,
            "contenido_caja": contenido_caja
        }

    # 3. Detectar formato del archivo
    fname_lower = filename.lower()
    filas_raw = []
    numero_factura_detectado = None
    proveedor_detectado = None
    fecha_factura_detectada = None
    formato_detectado = "DESCONOCIDO"

    # A. PDF (Facturas de Retail, Almacenes de Cadena, Alkosto, Éxito, Mayoristas, Ferreterías, Medicamentos y DIAN)
    es_pdf = fname_lower.endswith(".pdf") or file_bytes.strip().startswith(b"%PDF")
    if es_pdf:
        formato_detectado = "PDF_FACTURA_UNIVERSAL"
        try:
            import pdfplumber

            def _calibrar_triplete_universal(num_vals_list, dialecto="LATIN"):
                mejores = []
                for i_tot, tot_c in enumerate(num_vals_list):
                    tot_v = _smart_parse_num(tot_c, is_price=True, default=0.0, dialecto=dialecto)
                    if tot_v <= 0 or tot_v > 500000000:
                        continue
                    for i_cant, cant_c in enumerate(num_vals_list):
                        if i_cant == i_tot:
                            continue
                        cant_v = _smart_parse_num(cant_c, is_price=False, default=1.0, dialecto=dialecto)
                        if cant_v <= 0 or cant_v > 1000000:
                            continue
                        for i_p, p_c in enumerate(num_vals_list):
                            if i_p in [i_tot, i_cant]:
                                continue
                            p_v = _smart_parse_num(p_c, is_price=True, default=0.0, dialecto=dialecto)
                            if p_v <= 0 or p_v > 500000000:
                                continue
                            diff = abs(cant_v * p_v - tot_v)
                            if diff <= max(3.0, tot_v * 0.02):
                                score = 100
                                if cant_v != 1.0:
                                    score += 50
                                if p_v > 100:
                                    score += 30
                                mejores.append((score, -diff, cant_v, p_v, tot_v, tot_c, cant_c, p_c))

                if mejores:
                    mejores.sort(reverse=True)
                    best = mejores[0]
                    return best[2], best[3], best[4], [best[5], best[6], best[7]]
                return None, None, None, []

            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                texto_completo = ""
                for page in pdf.pages:
                    t = page.extract_text() or ""
                    texto_completo += t + "\n"

                dialecto_doc = _detectar_dialecto_numerico_texto(texto_completo)
                lineas_pdf = [l.strip() for l in texto_completo.splitlines() if l.strip()]

                # 1. Metadatos de Cabecera Universales (Bloque 1)
                for l in lineas_pdf[:35]:
                    l_str = l.strip()
                    # NIT
                    m_nit = re.search(r'\b(?:Nit|NIT|N\.I\.T\.)\s*[:.]?\s*(\d{7,11}(?:-\d)?)\b', l_str)
                    if m_nit and not proveedor_detectado:
                        proveedor_detectado = re.sub(r'\s*\|\s*.*$', '', l_str).strip()

                    # Razón Social / Proveedor
                    if not proveedor_detectado and any(term in l_str.upper() for term in ["S.A.S", "S.A.", "LTDA", "ALKOSTO", "EXITO", "SODIMAC", "HOMECENTER", "MAKRO", "COOPIDROGAS", "DISTRIBUIDORA", "DROGUERIA", "LABORATORIO", "LOINPRO", "AUDIFARMA", "DROMAYOR", "ETICOS", "DISTRACOM", "SOCIEDAD"]):
                        prov_cand = re.sub(r'\s*\|\s*.*$', '', l_str).strip()
                        prov_cand = re.sub(r'\b(?:FACTURA|ELECTRONICA|VENTA|NIT)\b.*$', '', prov_cand, flags=re.IGNORECASE).strip()
                        if len(prov_cand) >= 3:
                            proveedor_detectado = prov_cand

                    # Número de Factura
                    m_num = re.search(r'(?:N[°o]|Nro|Consecutivo|Numero|FEV|FEEV|F\.V\.|Factura(?:\s+Electronica)?(?:\s+de\s+Venta)?)\s*[:.]?\s*([A-Za-z0-9\-_]{3,25})', l_str, re.IGNORECASE)
                    if m_num and not numero_factura_detectado:
                        pos = m_num.group(1).strip()
                        if not re.match(r'^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}$', pos) and not re.match(r'^\d{1,2}[-/.]\d{1,2}[-/.]\d{4}$', pos):
                            if any(c.isdigit() for c in pos) and not any(k in pos.upper() for k in ["ELECTR", "VENTA", "ESTADO", "EMISION"]):
                                numero_factura_detectado = pos

                    # Fecha
                    m_fec = re.search(r'Fecha.*?(\d{4}[-/.]\d{1,2}[-/.]\d{1,2}|\d{1,2}[-/.]\d{1,2}[-/.]\d{4})', l_str, re.IGNORECASE)
                    if m_fec and not fecha_factura_detectada:
                        raw_f = m_fec.group(1)
                        m_f_iso = re.match(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$", raw_f)
                        if m_f_iso:
                            fecha_factura_detectada = f"{m_f_iso.group(1)}-{int(m_f_iso.group(2)):02d}-{int(m_f_iso.group(3)):02d}"
                        else:
                            m_f_lat = re.match(r"^(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})$", raw_f)
                            if m_f_lat:
                                fecha_factura_detectada = f"{m_f_lat.group(3)}-{int(m_f_lat.group(2)):02d}-{int(m_f_lat.group(1)):02d}"

                if not proveedor_detectado:
                    for l in lineas_pdf[:6]:
                        if not any(k in l.upper() for k in ['FACTURA', 'CLIENTE', 'FECHA', 'NIT', 'PEDIDO', 'HORA', 'CAJA', 'PEDIDO NO']):
                            if len(l) >= 3:
                                proveedor_detectado = l
                                break

                # ── ESTRATEGIA 1: Formato POS Retail / Tiquete Multilínea (Alkosto, Éxito, Cadenas) ────
                patron_pos_5num = re.compile(r'^(\d{1,4})\s+(\d{6,14})\s+(\d+(?:[.,]\d+)?)\s+(\d+(?:[.,]\d+)?)\s+(\d+(?:[.,]\d+)?)\s+(\d{1,3}(?:\.\d{3})*(?:,\d+)?|\d+(?:\.\d+)?)$')
                patron_pos_4num = re.compile(r'^(\d{1,4})\s+(\d{6,14})\s+(\d+(?:[.,]\d+)?)\s+(\d+(?:[.,]\d+)?)\s+(\d{1,3}(?:\.\d{3})*(?:,\d+)?|\d+(?:\.\d+)?)$')

                filas_extraidas = []
                i = 0
                while i < len(lineas_pdf):
                    l = lineas_pdf[i]
                    m5 = patron_pos_5num.match(l)
                    m4 = patron_pos_4num.match(l) if not m5 else None

                    if m5 or m4:
                        m = m5 if m5 else m4
                        num_item = int(m.group(1))
                        cod_barras = m.group(2)
                        if m5:
                            iva = float(m.group(3).replace(',', '.'))
                            ipo = float(m.group(4).replace(',', '.'))
                            cant = float(m.group(5).replace(',', '.'))
                            tot_raw = m.group(6).replace('.', '').replace(',', '.')
                        else:
                            iva = float(m.group(3).replace(',', '.'))
                            cant = float(m.group(4).replace(',', '.'))
                            tot_raw = m.group(5).replace('.', '').replace(',', '.')
                        total_s_desc = float(tot_raw)

                        nombre = ''
                        descuento_val = 0.0
                        i += 1
                        while i < len(lineas_pdf):
                            nxt = lineas_pdf[i]
                            if patron_pos_5num.match(nxt) or patron_pos_4num.match(nxt) or any(term in nxt.upper() for term in ['TOTAL LINEAS', 'SUBTOTAL', 'VALOR TOTAL', 'CUFE', 'RESPONSABLE I.V.A']):
                                break
                            m_desc = re.search(r'Descuento\s+(\d+(?:[.,]\d+)?)\s*%\s*(\d{1,3}(?:\.\d{3})*(?:,\d+)?|\d+(?:\.\d+)?)-?', nxt, re.IGNORECASE)
                            if m_desc:
                                d_raw = m_desc.group(2).replace('.', '').replace(',', '.')
                                descuento_val = float(d_raw)
                            else:
                                nombre = (nombre + ' ' + nxt).strip()
                            i += 1

                        tot_neto = max(0.0, total_s_desc - descuento_val)
                        costo_u = round(tot_neto / cant, 2) if cant > 0 else 0.0

                        filas_extraidas.append([
                            cod_barras,
                            cod_barras,
                            nombre,
                            cant,
                            costo_u,
                            iva,
                            0,
                            None,
                            None,
                            1
                        ])
                    else:
                        i += 1

                # ── ESTRATEGIA 2: Formato Mayorista / Distribución Estándar (Loinpro / Medicamentos) ────
                if not filas_extraidas:
                    idx_inicio = -1
                    for i_l, l in enumerate(lineas_pdf):
                        l_norm = re.sub(r'[^A-Za-z0-9\s.,/%+-]', ' ', l).upper()
                        if "ITEM" in l_norm and ("DIGO" in l_norm or "REF" in l_norm or "DESCRIP" in l_norm) and ("VR." in l_norm or "TOTAL" in l_norm or "CANT" in l_norm):
                            idx_inicio = i_l + 1
                            break

                    if idx_inicio != -1:
                        bloque_items = []
                        i_b = idx_inicio
                        while i_b < len(lineas_pdf):
                            l = lineas_pdf[i_b]
                            l_norm = re.sub(r'[^A-Za-z0-9\s.,/%+-]', ' ', l).upper()
                            if any(term in l_norm for term in ["TOTAL NRO. ITEMS", "TOTAL NRO ITEMS", "TOTAL ITEMS DOCUMENTO", "OBSERVACIONES", "SUBTOTAL", "VALOR EN LETRAS", "SON:", "TOTAL FACTURA", "CUFE"]):
                                break
                            bloque_items.append(l)
                            i_b += 1

                        patron_con_item = re.compile(r'^(\d{1,4})\s+([A-Za-z0-9\-]{4,25})\s+(.*)$')
                        items_unificados = []
                        item_actual = None
                        ultimo_item_num = 0

                        for l in bloque_items:
                            m1 = patron_con_item.match(l)
                            es_nuevo = False
                            cod = ""
                            resto = ""

                            if m1:
                                num = int(m1.group(1))
                                if num == ultimo_item_num + 1 or (ultimo_item_num == 0 and num in [1, 2]):
                                    es_nuevo = True
                                    ultimo_item_num = num
                                    cod = m1.group(2)
                                    resto = m1.group(3)

                            if es_nuevo:
                                if item_actual:
                                    items_unificados.append(item_actual)
                                item_actual = {"codigo": cod, "resto": resto, "lineas_extra": []}
                            else:
                                if item_actual:
                                    item_actual["lineas_extra"].append(l)

                        if item_actual:
                            items_unificados.append(item_actual)

                        patron_numeros_finales = re.compile(
                            r'^(?P<desc_y_lote>.*?)\s+'
                            r'(?:(?P<umedida>[A-Za-z]{2,5})\s+)?'
                            r'(?:(?P<iva>\d+(?:[.,]\d+)?)\s+)?'
                            r'(?P<cant>\d+(?:[.,]\d+)?)\s+'
                            r'(?P<vr_unit>\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?|\d+(?:[.,]\d+)?)\s+'
                            r'(?:(?P<dcto>\d+(?:[.,]\d+)?)\s+)?'
                            r'(?P<vr_total>\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})?|\d+(?:[.,]\d+)?)$'
                        )

                        for it in items_unificados:
                            codigo = it["codigo"]
                            resto = it["resto"]
                            extras = " ".join(it["lineas_extra"]).strip()

                            m_num = patron_numeros_finales.search(resto)
                            if m_num:
                                g = m_num.groupdict()
                                desc_y_lote = g["desc_y_lote"].strip()
                                if extras:
                                    desc_y_lote = desc_y_lote + " " + extras

                                cant = _smart_parse_num(g["cant"], is_price=False, default=1.0)
                                vr_u = _smart_parse_num(g["vr_unit"], is_price=True, default=0.0)
                                dcto = _smart_parse_num(g.get("dcto"), is_price=False, default=0.0)
                                iva = _smart_parse_num(g.get("iva"), is_price=False, default=0.0)
                                vr_tot = _smart_parse_num(g["vr_total"], is_price=True, default=0.0)

                                p_desc = _parse_pharma_description(desc_y_lote)
                                nom_limpio = p_desc["nombre"]
                                lote = p_desc["lote"]
                                venc = p_desc["vencimiento"]
                                cnt_caja = p_desc["contenido_caja"]

                                if vr_tot > 0 and cant > 0:
                                    costo_neto = round(vr_tot / cant, 2)
                                elif vr_u > 0:
                                    costo_neto = round(vr_u * (1.0 - dcto / 100.0), 2)
                                else:
                                    costo_neto = 0.0

                                filas_extraidas.append([
                                    codigo,
                                    codigo if len(codigo) >= 8 else "",
                                    nom_limpio,
                                    cant,
                                    costo_neto,
                                    iva,
                                    0,
                                    lote,
                                    venc,
                                    cnt_caja
                                ])

                # ── ESTRATEGIA 3: Extractor de Bloque Tabular Universal (Distracom, EDS, Ferreterías, Granel y Combustibles) ────
                if not filas_extraidas:
                    for page in pdf.pages:
                        # 3A: Tablas estructuradas nativas de pdfplumber
                        tables = page.extract_tables()
                        if tables:
                            for t in tables:
                                for row in t:
                                    clean_row = [str(c).strip() for c in row if c is not None and str(c).strip()]
                                    if len(clean_row) >= 4 and any(c.isdigit() for c in ''.join(clean_row)):
                                        r_norm = _normalizar(' '.join(clean_row))
                                        if any(k in r_norm for k in ['subtotal', 'base iva', 'iva:', 'total:', 'recargos']):
                                            continue
                                        
                                        num_vals = []
                                        nombre_t = ""
                                        for cell in clean_row:
                                            if re.match(r'^\$?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d+)?$', cell) or re.match(r'^\d+(?:[.,]\d+)?$', cell):
                                                num_vals.append(cell)
                                            elif len(cell) >= 3 and not any(k in _normalizar(cell) for k in ['galon', 'gln', 'und', 'unidad', 'kg', 'lt']):
                                                if not nombre_t:
                                                    nombre_t = cell
                                                else:
                                                    nombre_t += " " + cell

                                        cant_t, precio_t, total_t, used_nums = _calibrar_triplete_universal(num_vals)
                                        codigo_t = ""
                                        for rem in num_vals:
                                            if rem not in used_nums and (len(rem) >= 4 or (rem.isdigit() and int(rem) > 100)):
                                                codigo_t = rem
                                                break

                                        if nombre_t and cant_t:
                                            filas_extraidas.append([
                                                codigo_t or "000001",
                                                codigo_t if len(codigo_t) >= 8 else "",
                                                nombre_t,
                                                cant_t,
                                                precio_t or round(total_t / cant_t, 2),
                                                0.0,
                                                0,
                                                None,
                                                None,
                                                1
                                            ])
                            if filas_extraidas:
                                break

                        # 3B: Búsqueda visual por bloques de palabras X/Y
                        words = page.extract_words()
                        if not words:
                            continue

                        header_keys = ['item', 'ref', 'codigo', 'cod', 'producto', 'descripcion', 'estacion', 'valor', 'unit', 'cant', 'cantidad', 'total', 'precio', 'iva']
                        header_top = -1
                        for w in words:
                            wn = _normalizar(w['text'])
                            if wn in ['item', 'codigo', 'producto', 'descripcion', 'cant.', 'cant', 'cantidad', 'estacion', 'ref.']:
                                neighbors = [nw for nw in words if abs(nw['top'] - w['top']) < 10.0 and _normalizar(nw['text']) in header_keys]
                                if len(neighbors) >= 3:
                                    header_top = min(nw['top'] for nw in neighbors)
                                    break

                        if header_top == -1:
                            continue

                        end_top = page.height
                        for w in words:
                            if w['top'] > header_top + 15.0:
                                wn = _normalizar(w['text'])
                                if wn in ['observaciones', 'subtotal:', 'subtotal', 'cufe', 'total:', 'base'] or 'total items' in wn or 'items documento' in wn:
                                    if w['top'] < end_top:
                                        end_top = w['top']

                        data_words = [w for w in words if header_top + 12.0 <= w['top'] < end_top]
                        data_words = [w for w in data_words if not any(term in _normalizar(w['text']) for term in ['documento:', 'observaciones', 'panel:'])]

                        rows = []
                        for w in sorted(data_words, key=lambda x: (x['top'], x['x0'])):
                            placed = False
                            for r in rows:
                                if abs(w['top'] - r['avg_top']) < 15.0:
                                    r['words'].append(w)
                                    r['avg_top'] = sum(x['top'] for x in r['words']) / len(r['words'])
                                    placed = True
                                    break
                            if not placed:
                                rows.append({'avg_top': w['top'], 'words': [w]})

                        for r in rows:
                            r_words = sorted(r['words'], key=lambda w: w['x0'])
                            num_tokens = []
                            text_tokens = []
                            for w in r_words:
                                t = w['text']
                                if re.match(r'^\$?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d+)?$', t) or re.match(r'^\d+(?:[.,]\d+)?$', t):
                                    num_tokens.append({'word': w, 'val': _smart_parse_num(t, is_price=True, default=0.0), 'raw': t})
                                else:
                                    text_tokens.append(w)

                            num_raws = [nt['raw'] for nt in num_tokens]
                            cant_w, precio_w, total_w, used_w = _calibrar_triplete_universal(num_raws)

                            codigo = ""
                            for nt in num_tokens:
                                if nt['raw'] not in used_w and ((nt['raw'].startswith('0') and len(nt['raw']) >= 4) or len(nt['raw']) >= 8):
                                    codigo = nt['raw']
                                    break

                            prod_words = [tw['text'] for tw in text_tokens if _normalizar(tw['text']) not in ['cobo', 'vasquez', 'estacion', 'distracom', 'eds', 'unit.', 'valor']]
                            if not prod_words and text_tokens:
                                prod_words = [tw['text'] for tw in text_tokens]
                            nombre = ' '.join(prod_words).strip()

                            if _normalizar(nombre) in ['total items', 'total', 'subtotal', 'observaciones', 'panel']:
                                continue

                            if nombre and cant_w and precio_w:
                                filas_extraidas.append([
                                    codigo or "000001",
                                    codigo if len(codigo) >= 8 else "",
                                    nombre,
                                    cant_w,
                                    precio_w,
                                    0.0,
                                    0,
                                    None,
                                    None,
                                    1
                                ])

                if filas_extraidas:
                    filas_raw = [["codigo", "codigo_barras", "nombre", "cantidad", "costo_unitario", "iva_porcentaje", "precio_sugerido", "lote", "vencimiento", "contenido_caja"]] + filas_extraidas

        except Exception as e:
            import traceback
            traceback.print_exc()
            filas_raw = []

    # B. XML (DIAN Factura Electrónica UBL 2.1 / AttachedDocument)
    elif es_xml:
        formato_detectado = "XML_DIAN"
        try:
            texto = file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            texto = file_bytes.decode("latin1", errors="ignore")

        if "<AttachedDocument" in texto and "&lt;Invoice" in texto:
            import html
            sub_xml = html.unescape(texto)
            idx_inv = sub_xml.find("<Invoice")
            if idx_inv != -1:
                end_inv = sub_xml.find("</Invoice>") + len("</Invoice>")
                texto = sub_xml[idx_inv:end_inv]

        try:
            # Limpiar namespaces para búsqueda universal
            clean_xml = re.sub(r'\sxmlns(:\w+)?="[^"]+"', '', texto)
            clean_xml = re.sub(r'<(/?)\w+:', r'<\1', clean_xml)
            root = ET.fromstring(clean_xml)

            for elem in root.iter():
                t = elem.tag.lower()
                val = (elem.text or "").strip()
                if not val:
                    continue
                if t in ["id", "invoicenumber", "consecutivo"] and not numero_factura_detectado:
                    numero_factura_detectado = val
                elif t in ["name", "registrationname", "razonsocial"] and not proveedor_detectado:
                    proveedor_detectado = val

            filas_raw = [["codigo", "codigo_barras", "nombre", "cantidad", "costo_unitario", "iva_porcentaje", "precio_sugerido"]]
            for l in root.iter():
                if l.tag.lower() in ["invoiceline", "creditnoteline", "lineafactura", "detalle"]:
                    desc = ""
                    cant = 1.0
                    price = 0.0
                    iva = 0.0
                    code = ""
                    for elem in l.iter():
                        t = elem.tag.lower()
                        val = (elem.text or "").strip()
                        if not val:
                            continue
                        if t in ["description", "itemname", "productname"] and not desc:
                            desc = val
                        elif t == "name" and not desc:
                            desc = val
                        elif t in ["invoicedquantity", "quantity", "cantidad", "cant"]:
                            cant = _smart_parse_num(val, is_price=False, default=1.0)
                        elif t in ["priceamount", "price", "unitprice", "valorunitario", "preciocosto", "costo"]:
                            price = _smart_parse_num(val, is_price=True, default=0.0)
                        elif t in ["percent", "porcentaje", "iva", "taxpercent"]:
                            iva = _smart_parse_num(val, is_price=False, default=0.0)
                        elif t in ["id", "barcode", "ean", "ean13", "upc", "sku", "code"]:
                            if len(val) >= 8 or not code:
                                code = val

                    if desc or code:
                        filas_raw.append([code, code if len(code) >= 8 else "", desc or code, cant, price, iva, 0])
        except Exception:
            filas_raw = []

    # B. JSON
    elif fname_lower.endswith(".json") or file_bytes.strip().startswith(b"{") or file_bytes.strip().startswith(b"["):
        formato_detectado = "JSON"
        try:
            data = json.loads(file_bytes.decode("utf-8", errors="ignore"))
            items_list = data if isinstance(data, list) else data.get("lineas") or data.get("items") or data.get("productos") or []
            if isinstance(data, dict):
                numero_factura_detectado = data.get("numero_factura") or data.get("factura") or data.get("numero")
                proveedor_detectado = data.get("proveedor")
            filas_raw = [["codigo", "codigo_barras", "nombre", "cantidad", "costo_unitario", "iva_porcentaje", "precio_sugerido"]]
            for it in items_list:
                filas_raw.append([
                    it.get("codigo", ""),
                    it.get("codigo_barras") or it.get("ean", ""),
                    it.get("nombre") or it.get("descripcion", ""),
                    it.get("cantidad", 1),
                    it.get("costo_unitario") or it.get("costo", 0),
                    it.get("iva_porcentaje") or it.get("iva", 0),
                    it.get("precio_sugerido") or it.get("precio_venta", 0),
                ])
        except Exception:
            filas_raw = []

    # C. Excel (.xlsx, .xls)
    elif fname_lower.endswith((".xlsx", ".xls", ".xlsm")):
        formato_detectado = "EXCEL"
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    filas_raw.append(list(row))
        except Exception:
            filas_raw = []

    # D. Texto plano / Archivos Planos .DAT / .TXT / .CSV / .PRN / .TSV
    else:
        formato_detectado = "PLANO_DAT_TXT" if fname_lower.endswith(".dat") else "TEXTO_DELIMITADO"
        try:
            texto = file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            try:
                texto = file_bytes.decode("latin1", errors="ignore")
            except:
                texto = file_bytes.decode("cp1252", errors="ignore")

        lineas_texto = [l.strip() for l in texto.splitlines() if l.strip()]

        # Buscar encabezados de factura tipo plano (F|... o H|...)
        for l in lineas_texto[:5]:
            if l.startswith(("F|", "H|", "FAC|")):
                partes = l.split("|") if "|" in l else l.split(";")
                for p in partes:
                    p_limpio = p.strip()
                    if re.match(r"^[A-Z0-9\-]{4,20}$", p_limpio) and any(c.isdigit() for c in p_limpio):
                        if not numero_factura_detectado:
                            numero_factura_detectado = p_limpio
                    if any(term in p_limpio.upper() for term in ["DISTRIBUIDORA", "DROGUERIA", "PROVEEDOR", "COMERCIO"]):
                        proveedor_detectado = p_limpio

        # Sniff delimiter
        posibles_delims = ["|", ";", "\t", ","]
        conteos = {d: sum(l.count(d) for l in lineas_texto[:15]) for d in posibles_delims}
        mejor_delim = max(conteos, key=conteos.get)

        if conteos[mejor_delim] > 0:
            for l in lineas_texto:
                # Si la línea empieza por D| (detalle), remover prefijo
                l_procesar = l
                if l.startswith("D|") and mejor_delim == "|":
                    l_procesar = l[2:]
                elif l.startswith(("F|", "H|", "T|", "TOTAL|")):
                    continue # Saltar líneas de cabecera/totales

                fila_partes = [c.strip() for c in l_procesar.split(mejor_delim)]
                if any(c != "" for c in fila_partes):
                    filas_raw.append(fila_partes)
        else:
            # Líneas sin delimitador común (posicionales)
            for l in lineas_texto:
                tokens = [t.strip() for t in re.split(r"\s{2,}", l) if t.strip()]
                if len(tokens) >= 2:
                    filas_raw.append(tokens)

    if not filas_raw:
        return {
            "total_filas": 0,
            "encontrados": 0,
            "nuevos": 0,
            "margen_predeterminado": margen_def,
            "rubro": rubro,
            "formato_detectado": formato_detectado,
            "numero_factura_detectado": numero_factura_detectado,
            "proveedor_detectado": proveedor_detectado,
            "escala_precios_detectada": 1,
            "items": [],
        }

    # 3.1 Detección del Formato Plano Delimitado (18-25 columnas con Col 1=Cod, Col 5=Nombre, Col 11=Barras, Col 7=Costo Descuento, Col 19=Total)
    es_plano_delimitado = False
    for f in filas_raw[:10]:
        if len(f) >= 11:
            nom_c = str(f[4]).strip() if len(f) > 4 else ""
            bar_c = str(f[10]).strip() if len(f) > 10 else ""
            if len(nom_c) >= 3 and any(c.isalpha() for c in nom_c) and (len(bar_c) >= 8 or re.match(r"^\d{8,14}$", bar_c)):
                es_plano_delimitado = True
                break

    if es_plano_delimitado:
        formato_detectado = "PLANO_DAT_DELIMITADO"
        fecha_factura_detectada = None
        items = []

        for idx, f in enumerate(filas_raw):
            if len(f) < 5:
                continue

            cod_drogueria = str(f[0]).strip()

            # Columna 2 (f[1]): Fecha del pedido (formato AAAA-MM-DD o variantes)
            col_fecha = str(f[1]).strip() if len(f) > 1 else ""
            if col_fecha and not fecha_factura_detectada:
                m_f = re.match(r"^(\d{4})[-/.]?(\d{2})[-/.]?(\d{2})$", col_fecha)
                if m_f:
                    fecha_factura_detectada = f"{m_f.group(1)}-{m_f.group(2)}-{m_f.group(3)}"
                else:
                    m_f_inv = re.match(r"^(\d{2})[-/.](\d{2})[-/.](\d{4})$", col_fecha)
                    if m_f_inv:
                        fecha_factura_detectada = f"{m_f_inv.group(3)}-{m_f_inv.group(2)}-{m_f_inv.group(1)}"
                    elif len(col_fecha) >= 8 and any(c.isdigit() for c in col_fecha):
                        fecha_factura_detectada = col_fecha

            # Columna 3 (f[2]): Número de factura o pedido
            col_factura_pedido = str(f[2]).strip() if len(f) > 2 else ""
            if col_factura_pedido and not numero_factura_detectado and col_factura_pedido != "0":
                numero_factura_detectado = col_factura_pedido

            # Columna 4 (f[3]): Código del producto oficial
            cod_prod = str(f[3]).strip() if len(f) > 3 else ""

            # Columna 5 (f[4]): Descripción / Nombre del producto
            nombre_raw = str(f[4]).strip() if len(f) > 4 else ""
            if not nombre_raw:
                continue

            # Columna 6 (f[5]): Cantidad
            cantidad = max(1.0, _smart_parse_num(f[5], is_price=False, default=1.0) if len(f) > 5 else 1.0)
            # Columna 7 (f[6]): Costo descuento
            costo_desc = _smart_parse_num(f[6], is_price=True, default=0.0) if len(f) > 6 else 0.0
            # Columna 8 (f[7]): Costo real
            costo_real = _smart_parse_num(f[7], is_price=True, default=0.0) if len(f) > 7 else 0.0
            # Columna 11 (f[10]): Código de barras EAN
            barras_raw = str(f[10]).strip() if len(f) > 10 else ""
            # Columna 19 (f[18]): Total pagar
            total_pagar = _smart_parse_num(f[18], is_price=True, default=0.0) if len(f) > 18 else 0.0

            # Calcular costo unitario exacto considerando descuento y validación contra total
            costo_final = 0.0
            if cantidad > 0 and total_pagar > 0:
                costo_calc = total_pagar / cantidad
                if costo_desc > 0:
                    ratio = costo_desc / costo_calc
                    if 80 <= ratio <= 120:
                        costo_final = round(costo_calc, 2)
                    elif abs(costo_desc - costo_calc) < 1.0:
                        costo_final = round(costo_desc, 2)
                    else:
                        costo_final = round(costo_desc, 2)
                else:
                    costo_final = round(costo_calc, 2)
            elif costo_desc > 0:
                costo_final = round(costo_desc, 2)
            else:
                costo_final = round(costo_real, 2)

            es_obs, es_comb, factor_sug = _detectar_promocion_o_regalo(nombre_raw, costo_final)

            # Cruce con BD
            prod_encontrado: Optional[Producto] = None
            if barras_raw and _normalizar(barras_raw) in prods_by_barcode:
                prod_encontrado = prods_by_barcode[_normalizar(barras_raw)]
            elif cod_prod and _normalizar(cod_prod) in prods_by_codigo:
                prod_encontrado = prods_by_codigo[_normalizar(cod_prod)]
            elif nombre_raw and _normalizar(nombre_raw) in prods_by_nombre:
                prod_encontrado = prods_by_nombre[_normalizar(nombre_raw)]

            if prod_encontrado:
                p = prod_encontrado
                c_unit = costo_final if costo_final > 0 else float(p.precio_costo or 0)
                precio_sug = float(p.precio_venta or 0)
                if precio_sug <= 0 and c_unit > 0:
                    precio_sug = _redondear_precio(c_unit * (1 + margen_def / 100), modo_redondeo)
                ganancia = precio_sug - c_unit
                margen_pct = round((ganancia / c_unit) * 100, 2) if c_unit > 0 else margen_def

                items.append({
                    "index": idx,
                    "estado": "ENCONTRADO",
                    "producto_id": p.id,
                    "codigo": p.codigo,
                    "codigo_barras": p.codigo_barras or barras_raw or "",
                    "codigo_barras_blister": p.codigo_barras_blister or "",
                    "codigo_barras_unidad": p.codigo_barras_unidad or "",
                    "nombre": p.nombre,
                    "nombre_original": nombre_raw,
                    "principio_activo": p.principio_activo or "",
                    "laboratorio": p.laboratorio or "COOPIDROGAS",
                    "cantidad": cantidad,
                    "costo_unitario": c_unit,
                    "iva_porcentaje": float(p.iva_porcentaje or 0),
                    "precio_sugerido": precio_sug,
                    "porcentaje_ganancia": margen_pct,
                    "stock_actual_bd": float(p.stock_actual or 0),
                    "maneja_fracciones": p.maneja_fracciones,
                    "contenido_caja": p.contenido_caja or 1,
                    "contenido_blister": p.contenido_blister or 0,
                    "precio_caja": _redondear_precio(float(p.precio_caja or precio_sug), modo_redondeo),
                    "precio_blister": _redondear_precio(float(p.precio_blister or 0), modo_redondeo) if (float(p.precio_blister or 0) > 0) else 0.0,
                    "precio_unidad": _redondear_precio(float(p.precio_unidad or 0), modo_redondeo) if (float(p.precio_unidad or 0) > 0) else 0.0,
                    "es_obsequio_probable": es_obs,
                    "es_combo_probable": es_comb,
                    "factor_combo_sugerido": factor_sug,
                })
            else:
                precio_sug = _redondear_precio(costo_final * (1 + margen_def / 100), modo_redondeo)
                ganancia = precio_sug - costo_final
                margen_pct = round((ganancia / costo_final) * 100, 2) if costo_final > 0 else margen_def

                # Detección inteligente de unidades en el nombre (ej. 12 UND, X 100)
                m_qty = re.search(r'(?:X|\b)(\d+)\s*(?:UND|UNID|TAB|CAP|SOB|AMP|PZ)\b', nombre_raw, re.IGNORECASE)
                cnt_caja = int(m_qty.group(1)) if m_qty else 1
                p_caja_nuevo = precio_sug
                p_unidad_nuevo = _redondear_precio((p_caja_nuevo / cnt_caja) * 1.25, modo_redondeo) if cnt_caja > 1 else 0.0

                items.append({
                    "index": idx,
                    "estado": "NUEVO",
                    "producto_id": None,
                    "codigo": cod_prod or f"PROD-{idx + 1:04d}",
                    "codigo_barras": barras_raw,
                    "codigo_barras_blister": "",
                    "codigo_barras_unidad": "",
                    "nombre": nombre_raw,
                    "nombre_original": nombre_raw,
                    "principio_activo": "",
                    "laboratorio": "COOPIDROGAS",
                    "cantidad": cantidad,
                    "costo_unitario": costo_final,
                    "iva_porcentaje": 0,
                    "precio_sugerido": precio_sug,
                    "porcentaje_ganancia": margen_pct,
                    "stock_actual_bd": 0,
                    "maneja_fracciones": True if (rubro == "FARMACIA" or cnt_caja > 1) else False,
                    "contenido_caja": cnt_caja,
                    "contenido_blister": 0,
                    "precio_caja": p_caja_nuevo,
                    "precio_blister": 0.0,
                    "precio_unidad": p_unidad_nuevo,
                    "es_obsequio_probable": es_obs,
                    "es_combo_probable": es_comb,
                    "factor_combo_sugerido": factor_sug,
                })

        # Verificación de si la factura de Coopidrogas ya fue ingresada previamente
        factura_ya_registrada = False
        compra_previa_info = None
        if numero_factura_detectado:
            from app.models.inventario import Compra
            stmt_c = (
                select(Compra)
                .options(joinedload(Compra.proveedor), joinedload(Compra.lineas))
                .where(Compra.numero_factura_proveedor == numero_factura_detectado)
            )
            res_c = await db.execute(stmt_c)
            c_prev = res_c.scalars().unique().first()
            if c_prev:
                factura_ya_registrada = True
                compra_previa_info = {
                    "id": c_prev.id,
                    "numero": c_prev.numero,
                    "numero_factura_proveedor": c_prev.numero_factura_proveedor,
                    "fecha": c_prev.fecha.isoformat() if c_prev.fecha else None,
                    "total": float(c_prev.total or 0),
                    "proveedor_nombre": c_prev.proveedor.razon_social if c_prev.proveedor else "Coopidrogas",
                    "total_items": len(c_prev.lineas),
                }

        return {
            "total_filas": len(items),
            "encontrados": sum(1 for x in items if x["estado"] == "ENCONTRADO"),
            "nuevos": sum(1 for x in items if x["estado"] == "NUEVO"),
            "margen_predeterminado": margen_def,
            "rubro": rubro,
            "modo_redondeo": modo_redondeo,
            "formato_detectado": formato_detectado,
            "numero_factura_detectado": numero_factura_detectado,
            "proveedor_detectado": proveedor_detectado,
            "fecha_detectada": fecha_factura_detectada,
            "escala_precios_detectada": 1,
            "factura_ya_registrada": factura_ya_registrada,
            "compra_previa": compra_previa_info,
            "items": items,
        }

    # 4. Diccionario ampliado de sinónimos (incluyendo terminología Coopidrogas, Audifarma, AXON, DIAN)
    sinonimos = {
        "codigo_barras": [
            "codigo_barras", "cod_barras", "ean", "barcode", "barras", "upc", "gtin",
            "cod_barras_caja", "codigo_de_barras", "codigo barras", "c_barras", "ean13",
            "codbarras", "c_ean", "plu", "barcode_caja", "cbarras"
        ],
        "codigo_barras_blister": ["cod_barras_blister", "codigo_barras_blister", "barcode_blister", "ean_blister"],
        "codigo_barras_unidad": ["cod_barras_unidad", "codigo_barras_unidad", "barcode_unidad", "ean_unidad"],
        "codigo": [
            "codigo", "cod", "ref", "referencia", "sku", "codigo_articulo", "item", "codigo_producto",
            "cod_articulo", "cod_art", "cod_coopidrogas", "cod_cop", "cod_prov", "cod_distribuidor"
        ],
        "nombre": [
            "nombre", "descripcion", "producto", "articulo", "detalle", "descripcion_articulo", "nombre_producto",
            "desc_articulo", "nom_articulo", "item_description", "concepto", "descripcion_producto"
        ],
        "cantidad": [
            "cantidad", "cant", "unidades", "qty", "cajas", "bultos", "cant_comprada", "cant_facturada",
            "cant_enviada", "cant_despachada", "unid_facturadas", "cant_pedida"
        ],
        "costo_unitario": [
            "costo", "costo_unitario", "precio_costo", "costo_unit", "p_costo", "vlr_unitario", "vr_unit",
            "valor_unitario", "precio_compra", "costo_neto", "vr_neto_unit", "precio_unitario", "costo_unitario_neto"
        ],
        "iva_porcentaje": ["iva", "iva_porcentaje", "pct_iva", "tarifa_iva", "porcentaje_iva", "%iva", "tarifa"],
        "precio_sugerido": [
            "precio_venta", "precio_sugerido", "pvp", "p_venta", "precio_publico", "venta", "p_publico",
            "pvp_sugerido", "vr_publico", "precio_sugerido_venta"
        ],
        "principio_activo": ["principio_activo", "sustancia", "generico", "molecula", "droga", "droga_generica"],
        "laboratorio": ["laboratorio", "marca", "fabricante", "proveedor_marca", "lab", "casa_comercial"],
        "lote": ["lote", "batch", "lot", "nro_lote", "num_lote", "no_lote"],
        "vencimiento": ["vencimiento", "vence", "fecha_vencimiento", "fecha_venc", "f_vto", "f_venc", "exp", "expiry", "vto"],
        "contenido_caja": ["contenido_caja", "fraccion", "unidades_caja", "cant_caja", "contenido", "unidades por caja", "unid_caja"],
        "contenido_blister": ["contenido_blister", "unidades_blister", "blister_caja", "unidades por blister", "unid_blister"],
        "precio_caja": ["precio_caja", "pvp_caja"],
        "precio_blister": ["precio_blister", "pvp_blister"],
        "precio_unidad": ["precio_unidad", "pvp_unidad"],
    }

    mejor_idx_header = 0
    max_coincidencias = 0
    mapa_cols = {}

    for r_idx in range(min(12, len(filas_raw))):
        fila = filas_raw[r_idx]
        temp_map = {}
        matches = 0
        for c_idx, cell in enumerate(fila):
            val_norm = _normalizar(cell)
            if not val_norm:
                continue
            for col_estandar, lista_sin in sinonimos.items():
                if col_estandar not in temp_map:
                    if any(sin in val_norm for sin in lista_sin):
                        temp_map[col_estandar] = c_idx
                        matches += 1
                        break
        if matches > max_coincidencias:
            max_coincidencias = matches
            mejor_idx_header = r_idx
            mapa_cols = temp_map

    # Si no se detectó cabecera clara (ej: archivo Coopidrogas .DAT posicional directo)
    tiene_cabecera_clara = max_coincidencias >= 2
    filas_datos = filas_raw[mejor_idx_header + 1:] if tiene_cabecera_clara else filas_raw

    # Heurística posicional especializada para Coopidrogas cuando no hay encabezados
    if not tiene_cabecera_clara and filas_datos:
        primera_fila = filas_datos[0]
        n_cols = len(primera_fila)
        mapa_cols = {}

        # Coopidrogas estándar: [EAN, COD_INT, NOMBRE, LAB, CANT, BONIF, COSTO, IVA, PVP, SUSTANCIA, ...]
        if n_cols >= 7 and re.match(r"^\d{12,14}$", str(primera_fila[0]).strip()):
            mapa_cols["codigo_barras"] = 0
            mapa_cols["codigo"] = 1
            mapa_cols["nombre"] = 2
            c3 = str(primera_fila[3]).strip()
            if any(c.isalpha() for c in c3):
                mapa_cols["laboratorio"] = 3
                mapa_cols["cantidad"] = 4
                mapa_cols["costo_unitario"] = 6 if n_cols > 6 else 5
                if n_cols > 7:
                    mapa_cols["iva_porcentaje"] = 7
                if n_cols > 8:
                    mapa_cols["precio_sugerido"] = 8
                if n_cols > 9 and any(c.isalpha() for c in str(primera_fila[9])):
                    mapa_cols["principio_activo"] = 9
            else:
                mapa_cols["cantidad"] = 3
                mapa_cols["costo_unitario"] = 4
                if n_cols > 5:
                    mapa_cols["iva_porcentaje"] = 5
                if n_cols > 6:
                    mapa_cols["precio_sugerido"] = 6
        else:
            # Heurística posicional genérica
            for c_i, val in enumerate(primera_fila):
                s_val = str(val).strip()
                if re.match(r"^\d{12,14}$", s_val) and "codigo_barras" not in mapa_cols:
                    mapa_cols["codigo_barras"] = c_i
                elif re.match(r"^[A-Za-z0-9\-]{3,8}$", s_val) and "codigo" not in mapa_cols:
                    mapa_cols["codigo"] = c_i
                elif len(s_val) > 8 and any(c.isalpha() for c in s_val) and "nombre" not in mapa_cols:
                    mapa_cols["nombre"] = c_i

            if "nombre" in mapa_cols:
                restantes = [i for i in range(n_cols) if i not in mapa_cols.values()]
                if restantes:
                    mapa_cols["cantidad"] = restantes[0]
                if len(restantes) > 1:
                    mapa_cols["costo_unitario"] = restantes[1]
                if len(restantes) > 2:
                    mapa_cols["iva_porcentaje"] = restantes[2]
                if len(restantes) > 3:
                    mapa_cols["precio_sugerido"] = restantes[3]

    items = []

    for idx, f in enumerate(filas_datos):
        def _get(col_name, default=None):
            if col_name in mapa_cols:
                col_i = mapa_cols[col_name]
                if col_i < len(f) and f[col_i] is not None:
                    val = str(f[col_i]).strip()
                    if val != "" and val.lower() != "none":
                        return val
            return default

        nombre_raw = _get("nombre", "")
        codigo_raw = _get("codigo", "")
        barras_raw = _get("codigo_barras", "")
        barras_blister_raw = _get("codigo_barras_blister", "")
        barras_unidad_raw = _get("codigo_barras_unidad", "")
        lote_raw = _get("lote", None)
        vencimiento_raw = _get("vencimiento", None)

        # Si no hay ni nombre ni código ni barras, ignorar fila vacía
        if not nombre_raw and not codigo_raw and not barras_raw:
            continue

        cant_val_raw = _smart_parse_num(_get("cantidad", 1), is_price=False, default=1.0)
        cantidad = cant_val_raw if cant_val_raw > 0 else 1.0
        costo_unit = _smart_parse_num(_get("costo_unitario", 0), is_price=True, default=0.0)
        iva_pct = _smart_parse_num(_get("iva_porcentaje", 0), is_price=False, default=0.0)
        precio_sug_excel = _smart_parse_num(_get("precio_sugerido", None), is_price=True, default=None)

        principio_activo = _get("principio_activo", "")
        laboratorio = _get("laboratorio", "")
        contenido_caja = int(_smart_parse_num(_get("contenido_caja", 1), is_price=False, default=1))
        contenido_blister = int(_smart_parse_num(_get("contenido_blister", 0), is_price=False, default=0))
        precio_caja = _smart_parse_num(_get("precio_caja", 0), is_price=True, default=0.0)
        precio_blister = _smart_parse_num(_get("precio_blister", 0), is_price=True, default=0.0)
        precio_unidad = _smart_parse_num(_get("precio_unidad", 0), is_price=True, default=0.0)

        # 5. Cruce inteligente con BD
        prod_encontrado: Optional[Producto] = None
        if barras_raw and _normalizar(barras_raw) in prods_by_barcode:
            prod_encontrado = prods_by_barcode[_normalizar(barras_raw)]
        elif codigo_raw and _normalizar(codigo_raw) in prods_by_codigo:
            prod_encontrado = prods_by_codigo[_normalizar(codigo_raw)]
        elif nombre_raw and _normalizar(nombre_raw) in prods_by_nombre:
            prod_encontrado = prods_by_nombre[_normalizar(nombre_raw)]

        es_obs, es_comb, factor_sug = _detectar_promocion_o_regalo(nombre_raw, costo_unit)

        if prod_encontrado:
            p = prod_encontrado
            costo_factura = costo_unit if costo_unit > 0 else float(p.precio_costo or 0)
            costo_ant = float(p.precio_costo or 0)
            stock_ant = float(p.stock_actual or 0)
            cant_nueva = cantidad

            # Cálculo de Costo Promedio Ponderado vs Costo Más Alto vs Último Costo
            if stock_ant > 0 and costo_ant > 0 and costo_factura > 0:
                cpp = round(((stock_ant * costo_ant) + (cant_nueva * costo_factura)) / (stock_ant + cant_nueva), 2)
                costo_max = max(costo_ant, costo_factura)
                costo_ult = costo_factura
                cambio_costo = abs(costo_factura - costo_ant) > 0.01
            else:
                cpp = costo_factura
                costo_max = costo_factura
                costo_ult = costo_factura
                cambio_costo = False

            costo_final = cpp
            precio_final = precio_sug_excel if (precio_sug_excel and precio_sug_excel > 0) else float(p.precio_venta or 0)
            if precio_final <= 0 and costo_final > 0:
                precio_final = _redondear_precio(costo_final * (1 + margen_def / 100), modo_redondeo)

            ganancia = precio_final - costo_final
            margen_pct = round((ganancia / costo_final) * 100, 2) if costo_final > 0 else margen_def

            items.append({
                "index": idx,
                "estado": "ENCONTRADO",
                "producto_id": p.id,
                "codigo": p.codigo,
                "codigo_barras": p.codigo_barras or barras_raw or "",
                "codigo_barras_blister": p.codigo_barras_blister or barras_blister_raw or "",
                "codigo_barras_unidad": p.codigo_barras_unidad or barras_unidad_raw or "",
                "nombre": p.nombre,
                "nombre_original": nombre_raw,
                "principio_activo": p.principio_activo or principio_activo or "",
                "laboratorio": p.laboratorio or laboratorio or "",
                "lote": lote_raw,
                "vencimiento": vencimiento_raw,
                "cantidad": cantidad,
                "costo_unitario": costo_factura,
                "costo_factura": costo_factura,
                "costo_calculado_producto": cpp,
                "costo_anterior_bd": costo_ant,
                "costo_promedio_ponderado": cpp,
                "costo_mas_alto": costo_max,
                "costo_ultimo": costo_ult,
                "estrategia_costo": "PROMEDIO_PONDERADO",
                "cambio_costo_detectado": cambio_costo,
                "stock_actual_bd": stock_ant,
                "iva_porcentaje": iva_pct if iva_pct > 0 else float(p.iva_porcentaje or 0),
                "precio_sugerido": precio_final,
                "porcentaje_ganancia": margen_pct,
                "maneja_fracciones": p.maneja_fracciones or False,
                "contenido_caja": p.contenido_caja or 1,
                "contenido_blister": p.contenido_blister or 0,
                "precio_caja": _redondear_precio(float(p.precio_caja or precio_final), modo_redondeo),
                "precio_blister": _redondear_precio(float(p.precio_blister or 0), modo_redondeo) if (float(p.precio_blister or 0) > 0) else 0.0,
                "precio_unidad": _redondear_precio(float(p.precio_unidad or 0), modo_redondeo) if (float(p.precio_unidad or 0) > 0) else 0.0,
                "es_obsequio_probable": es_obs,
                "es_combo_probable": es_comb,
                "factor_combo_sugerido": factor_sug,
            })
        else:
            # Producto Nuevo
            costo_final = costo_unit
            precio_final = precio_sug_excel if (precio_sug_excel and precio_sug_excel > 0) else _redondear_precio(costo_final * (1 + margen_def / 100), modo_redondeo)
            ganancia = precio_final - costo_final
            margen_pct = round((ganancia / costo_final) * 100, 2) if costo_final > 0 else margen_def

            p_caja_nuevo = precio_final
            p_blister_nuevo = _redondear_precio(precio_blister, modo_redondeo) if precio_blister > 0 else 0.0
            p_unidad_nuevo = _redondear_precio(precio_unidad, modo_redondeo) if precio_unidad > 0 else 0.0

            if p_unidad_nuevo <= 0 and contenido_caja > 1:
                p_unidad_nuevo = _redondear_precio((p_caja_nuevo / contenido_caja) * 1.25, modo_redondeo)

            items.append({
                "index": idx,
                "estado": "NUEVO",
                "producto_id": None,
                "codigo": codigo_raw or f"PROD-{idx + 1:04d}",
                "codigo_barras": barras_raw or "",
                "codigo_barras_blister": barras_blister_raw or "",
                "codigo_barras_unidad": barras_unidad_raw or "",
                "nombre": nombre_raw or f"Producto sin nombre {idx + 1}",
                "nombre_original": nombre_raw or "",
                "principio_activo": principio_activo,
                "laboratorio": laboratorio,
                "lote": lote_raw,
                "vencimiento": vencimiento_raw,
                "cantidad": cantidad,
                "costo_unitario": costo_final,
                "costo_factura": costo_final,
                "costo_calculado_producto": costo_final,
                "costo_anterior_bd": 0,
                "costo_promedio_ponderado": costo_final,
                "costo_mas_alto": costo_final,
                "costo_ultimo": costo_final,
                "estrategia_costo": "PROMEDIO_PONDERADO",
                "iva_porcentaje": iva_pct,
                "precio_sugerido": precio_final,
                "porcentaje_ganancia": margen_pct,
                "stock_actual_bd": 0,
                "maneja_fracciones": True if (rubro == "FARMACIA" or contenido_caja > 1) else False,
                "contenido_caja": contenido_caja,
                "contenido_blister": contenido_blister,
                "precio_caja": p_caja_nuevo,
                "precio_blister": p_blister_nuevo,
                "precio_unidad": p_unidad_nuevo,
                "es_obsequio_probable": es_obs,
                "es_combo_probable": es_comb,
                "factor_combo_sugerido": factor_sug,
            })

    # 6. Auto-detección de escala de decimales implícitos (ej. 1250000 para $12.500)
    escala_detectada = 1
    ratios_100 = 0
    comparables = 0
    for it in items:
        c = it["costo_unitario"]
        if it["producto_id"] and it["producto_id"] in prods_by_id:
            p_bd = prods_by_id[it["producto_id"]]
            costo_bd = float(p_bd.precio_costo or 0)
            if costo_bd > 0 and c > 0:
                comparables += 1
                if 70 <= (c / costo_bd) <= 130:
                    ratios_100 += 1

    if comparables > 0 and (ratios_100 / comparables) >= 0.5:
        escala_detectada = 100
    elif not comparables and items:
        costos_validos = [x["costo_unitario"] for x in items if x["costo_unitario"] > 0]
        if costos_validos:
            mediana = sorted(costos_validos)[len(costos_validos) // 2]
            if mediana >= 300000 and all(c % 10 == 0 for c in costos_validos):
                escala_detectada = 100

    if escala_detectada > 1:
        for it in items:
            it["costo_unitario"] = round(it["costo_unitario"] / escala_detectada, 2)
            if it["precio_sugerido"]:
                it["precio_sugerido"] = _redondear_precio(it["precio_sugerido"] / escala_detectada, modo_redondeo)
            if it["precio_caja"]:
                it["precio_caja"] = _redondear_precio(it["precio_caja"] / escala_detectada, modo_redondeo)
            if it["precio_blister"]:
                it["precio_blister"] = _redondear_precio(it["precio_blister"] / escala_detectada, modo_redondeo)
            if it["precio_unidad"]:
                it["precio_unidad"] = _redondear_precio(it["precio_unidad"] / escala_detectada, modo_redondeo)
            c_fin = it["costo_unitario"]
            p_fin = it["precio_sugerido"]
            if c_fin > 0 and p_fin > 0:
                it["porcentaje_ganancia"] = round(((p_fin - c_fin) / c_fin) * 100, 2)

    # Verificación de si la factura ya fue ingresada previamente
    factura_ya_registrada = False
    compra_previa_info = None
    if numero_factura_detectado:
        from app.models.inventario import Compra
        stmt_c = (
            select(Compra)
            .options(joinedload(Compra.proveedor), joinedload(Compra.lineas))
            .where(Compra.numero_factura_proveedor == numero_factura_detectado)
        )
        res_c = await db.execute(stmt_c)
        c_prev = res_c.scalars().unique().first()
        if c_prev:
            factura_ya_registrada = True
            compra_previa_info = {
                "id": c_prev.id,
                "numero": c_prev.numero,
                "numero_factura_proveedor": c_prev.numero_factura_proveedor,
                "fecha": c_prev.fecha.isoformat() if c_prev.fecha else None,
                "total": float(c_prev.total or 0),
                "proveedor_nombre": c_prev.proveedor.razon_social if c_prev.proveedor else "Proveedor",
                "total_items": len(c_prev.lineas),
            }

    return {
        "total_filas": len(items),
        "encontrados": sum(1 for x in items if x["estado"] == "ENCONTRADO"),
        "nuevos": sum(1 for x in items if x["estado"] == "NUEVO"),
        "margen_predeterminado": margen_def,
        "rubro": rubro,
        "modo_redondeo": modo_redondeo,
        "formato_detectado": formato_detectado,
        "numero_factura_detectado": numero_factura_detectado,
        "proveedor_detectado": proveedor_detectado,
        "fecha_detectada": fecha_factura_detectada if 'fecha_factura_detectada' in locals() else None,
        "escala_precios_detectada": escala_detectada,
        "factura_ya_registrada": factura_ya_registrada,
        "compra_previa": compra_previa_info,
        "items": items,
    }